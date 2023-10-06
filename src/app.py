from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_mysqldb import MySQL
from flask_wtf.csrf import CSRFProtect
from flask_login import LoginManager, current_user, login_user, logout_user, login_required
from openpyxl import Workbook
import io
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import requests
import bcrypt


from config import config

# Models:
from models.ModelUsers import ModelUser

# Entities:
from models.entities.User import User

app = Flask(__name__)

# ( CAMBIAR A True PARA MANTENIMIENTO,  False FUNCIONAMIENTO NORMAL)
app.config['maintenance_mode'] = False

csrf = CSRFProtect()
db = MySQL(app)
login_manager_app = LoginManager(app)


@login_manager_app.user_loader
def load_user(id):
    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/admin_user/loginWEB.php')
    if response.status_code == 200:
        users = response.json()

        for user in users:
            if str(user['id']) == str(id):
                return User(user['id'], user['username'], user['password'], user['fullname'], user['rol'])

    return None


@app.route('/')
def index():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    else:
        return redirect(url_for('home'))


# **************************************************************************************************************************
# *************************************************************PP VAL EXCEL*************************************************************
# **************************************************************************************************************************


@app.route('/generate_excelPPVal', methods=['POST'])
@login_required
def generate_excelPPVal():
    # Retrieve form data from the request
    lote = request.form['lote']
    fecha_ingreso = request.form['fecha_ingreso']
    fecha = request.form['fecha']
    proveedor = request.form['proveedor']
    mp_gorgojo = request.form['mp_gorgojo']
    mp_partida = request.form['mp_partida']
    mp_fisicos = request.form['mp_fisicos']
    mp_quimicos = request.form['mp_quimicos']
    mp_biologico = request.form['mp_biologico']
    cantidadesxlargoPP = request.form['cantidadesxlargoPP']
    mp_validado_por = request.form['mp_validado_por']

    diferenciam3PP = request.form['diferenciam3PP']
    diferenciapulgPP = request.form['diferenciapulgPP']
    diferenciacantidadPP = request.form['diferenciacantidadPP']

    m3PP = request.form['m3PP']
    pulgPP = request.form['pulgPP']
    cantidadtotalPP = request.form['cantidadtotalPP']
    m3_totalPP = request.form['m3_totalPP']
    pulg_totalPP = request.form['pulg_totalPP']
    unidades_totalPP = request.form['unidades_totalPP']

    calibrePP = request.form['calibrePP']
    anchoPP = request.form['anchoPP']
    largoPP = request.form['largoPP']
    cantidadPP = request.form['cantidadPP']
    diff_unds = request.form['diff_unds']

    # Add more fields as needed

    # Create a new workbook
    wb = Workbook()
    ws = wb.active

    # LINEAS DE CUADRICULA DESACTIVADO
    sheet = wb.active
    sheet.sheet_view.showGridLines = False

    ancho_id = 3
    columna = ws.column_dimensions['A']
    columna.width = ancho_id
    columna = ws.column_dimensions['C']
    columna.width = ancho_id
    columna = ws.column_dimensions['N']
    columna.width = ancho_id

    ancho_diff1 = 2
    columna = ws.column_dimensions['H']
    columna.width = ancho_diff1
    columna = ws.column_dimensions['M']
    columna.width = ancho_diff1

    ancho_diff1 = 5
    columna = ws.column_dimensions['J']
    columna.width = ancho_diff1
    columna = ws.column_dimensions['K']
    columna.width = ancho_diff1

    # Write the form data to the worksheet
    ws.merge_cells('C1:Q4')  # Merge cells E1 to J2
    ws['C1'] = 'FORMATO INFORME LIBERADO PP'

    # Set alignment and formatting for the merged cell
    merged_cell = ws['C1']
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = Font(bold=True, size=14)

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['C1:Q4']:
        for cell in row:
            cell.border = border

    ws['R1'] = 'Codigo'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R1'].border = border

    ws['S1'] = 'FSAA-VDT-01'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S1'].border = border

    ws['R2'] = 'Vigencia'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R2'].border = border

    ws['S2'] = '21/06/2022'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S2'].border = border

    ws['R3'] = 'Versión'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R3'].border = border

    ws['S3'] = '2'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S3'].border = border

    ws['R4'] = 'Pag'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R4'].border = border

    ws['S4'] = '1 de 1'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S4'].border = border

    font_size = 8
    cell_range = ['S1', 'R1', 'S2', 'R2', 'S3', 'R3', 'S4', 'R4']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)

    ws.merge_cells('C5:D5')
    ws['C5'] = 'LOTE:'
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    for col in range(3, 5):  # Columnas C y D
        for row in range(5, 6):  # Fila 5
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = Font(size=13)

    # Merge y estilo de la celda E5
    ws.merge_cells('E5:F5')
    ws['E5'] = lote
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    for col in range(5, 7):  # Columnas E y F
        for row in range(5, 6):  # Fila 5
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = Font(size=13)

    # Set alignment for the cell containing the lote value
    ws['B1'].alignment = Alignment(horizontal='left')

    # Write other form data to the worksheet
    ws.merge_cells('L7:O7')
    ws['L7'] = 'Fecha Ingreso:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['L7:O7']:
        for cell in row:
            cell.border = border

    ws['P7'] = fecha_ingreso
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    ws['P7'].border = border
    for row in ws['O7:P7']:
        for cell in row:
            cell.border = border

    ws.merge_cells('L8:O8')
    ws['L8'] = 'Fecha Validación:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['L8:O8']:
        for cell in row:
            cell.border = border

    ws['P8'] = fecha
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    ws['P8'].border = border
    for row in ws['O8:P8']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E7:F7')
    ws['E7'] = 'Proveedor:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E7:F7']:
        for cell in row:
            cell.border = border

    ws.merge_cells('G7:J7')
    ws['G7'] = proveedor
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['G7:J7']:
        for cell in row:
            cell.border = border


# *************** CRITERIO DE ACEPTACION Y RECHAZO ******************
    ws.merge_cells('A11:F11')
    ws['A11'] = 'CRITERIO DE ACEPTACION Y RECHAZO'
    ws['A11'].alignment = Alignment(horizontal='center')
    ws['A11'].font = Font(bold=True, size=8)

    # Crear el estilo de borde
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A11:F11']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A12:F12')
    ws['A12'] = 'La madera cumple con las medidas vs remisión?'
    # Crear el estilo de borde
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A12:F12']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A13:F13')
    ws['A13'] = 'La madera cumple con las cantidades vs remisión?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A13:F13']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A14:F14')
    ws['A14'] = 'La madera presenta gorgojo?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A14:F14']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A15:F15')
    ws['A15'] = 'La madera esta partida o quebrada?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A15:F15']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A16:F16')
    ws['A16'] = 'La madera tiene agentes contaminantes físicos?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A16:F16']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A17:F17')
    ws['A17'] = 'La madera tiene agentes contaminantes químicos?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A17:F17']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A18:F18')
    ws['A18'] = 'La madera tiene agentes contaminantes biológicos?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A18:F18']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A19:F19')
    ws['A19'] = 'La madera tiene salvoconducto de movilización?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A19:F19']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A20:F20')
    ws['A20'] = 'Requiere informe no conformidad para el proveedor?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A20:F20']:
        for cell in row:
            cell.border = border

    font_size = 8
    cell_range = ['A12', 'A13', 'A14', 'A15',
                  'A16', 'A17', 'A18', 'A19', 'A20']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)

    ws.merge_cells('G11:L11')
    ws['G11'] = 'Marque con una (X)'
    ws['G11'].alignment = Alignment(horizontal='center')
    ws['G11'].font = Font(bold=True, size=8)

    # Crear el estilo de borde
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['G11:L11']:
        for cell in row:
            cell.border = border

    ws['G12'] = 'Si'
    ws['G12'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='808080'))
    ws['G12'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H12'].border = border
    ws.merge_cells('H12:I12')

    ws['G13'] = 'Si'
    ws['G13'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G13'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H13'].border = border
    ws.merge_cells('H13:I13')

    ws['G14'] = 'Si'
    ws['G14'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G14'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H14'].border = border
    ws.merge_cells('H14:I14')

    ws['G15'] = 'Si'
    ws['G15'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G15'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H15'].border = border
    ws.merge_cells('H15:I15')

    ws['G16'] = 'Si'
    ws['G16'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G16'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H16'].border = border
    ws.merge_cells('H16:I16')

    ws['G17'] = 'Si'
    ws['G17'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G17'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H17'].border = border
    ws.merge_cells('H17:I17')

    ws['G18'] = 'Si'
    ws['G18'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G18'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H12'].border = border
    ws.merge_cells('H18:I18')

    ws['G19'] = 'Si'
    ws['G19'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G19'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H19'].border = border
    ws.merge_cells('H19:I19')

    ws['G20'] = 'Si'
    ws['G20'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='000000'))
    ws['G20'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='000000'))

    ws['H20'].border = border
    ws.merge_cells('H20:I20')

    # _____________NO______________
    ws['J12'] = 'No'
    ws['J12'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J12'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K12'].border = border
    ws.merge_cells('K12:L12')

    ws['J13'] = 'No'
    ws['J13'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J13'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K13'].border = border
    ws.merge_cells('K13:L13')

    ws['J14'] = 'No'
    ws['J14'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J14'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K14'].border = border
    ws.merge_cells('K14:L14')

    ws['J15'] = 'No'
    ws['J15'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J15'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K15'].border = border
    ws.merge_cells('K15:L15')

    ws['J16'] = 'No'
    ws['J16'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J16'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K16'].border = border
    ws.merge_cells('K16:L16')

    ws['J17'] = 'No'
    ws['J17'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J17'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K17'].border = border
    ws.merge_cells('K17:L17')

    ws['J18'] = 'No'
    ws['J18'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J18'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K18'].border = border
    ws.merge_cells('K18:L18')

    ws['J19'] = 'No'
    ws['J19'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J19'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K19'].border = border
    ws.merge_cells('K19:L19')

    ws['J20'] = 'No'
    ws['J20'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='000000'))
    ws['J20'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='000000'))

    ws['K20'].border = border
    ws.merge_cells('K20:L20')

    font_size = 8
    cell_range = ['G12', 'G13', 'G14', 'G15', 'G16', 'G17', 'G18', 'G19',
                  'G20', 'J12', 'J13', 'J14', 'J15', 'J16', 'J17', 'J18', 'J19', 'J20']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)

    ws.merge_cells('M11:S11')
    ws['M11'] = 'OBSERVACIONES'
    ws['M11'].alignment = Alignment(horizontal='center')
    ws['M11'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in ws['M11:S11']:
        for cell in row:
            cell.border = border

    # _____COMETARIOS______

    ws.merge_cells('M12:S12')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M12:S12']:
        for cell in row:
            cell.border = border

    ws.merge_cells('M13:S13')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M13:S13']:
        for cell in row:
            cell.border = border

    ws['M14'] = mp_gorgojo
    ws.merge_cells('M14:S14')

    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M14:S14']:
        for cell in row:
            cell.border = border

    ws['M15'] = mp_partida
    ws.merge_cells('M15:S15')

    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M15:S15']:
        for cell in row:
            cell.border = border

    ws['M16'] = mp_fisicos
    ws.merge_cells('M16:S16')

    ws.merge_cells('M16:S16')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M16:S16']:
        for cell in row:
            cell.border = border

    ws['M17'] = mp_quimicos
    ws.merge_cells('M17:S17')

    ws.merge_cells('M17:S17')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M17:S17']:
        for cell in row:
            cell.border = border

    ws['M18'] = mp_biologico
    ws.merge_cells('M18:S18')

    ws.merge_cells('M18:S18')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M18:S18']:
        for cell in row:
            cell.border = border

    ws.merge_cells('M19:S19')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M19:S19']:
        for cell in row:
            cell.border = border

    ws.merge_cells('M20:S20')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in ws['M20:S20']:
        for cell in row:
            cell.border = border

    font_size = 8
    cell_range = ['M14', 'M15', 'M16', 'M17', 'M18']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)
# ______________________________________________________________

    ws.merge_cells('A22:S22')
    ws['A22'] = 'PONDERACION'
    ws['A22'].alignment = Alignment(horizontal='center')
    ws['A22'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in ws['A22:S22']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A23:D24')
    ws['A23'] = 'Agentes Fisícos'
    ws['A23'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A23'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A23:D24']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E23:S24')
    ws['E23'] = 'Los peligros físicos constan de objetos extraños presentes en los alimentos, pudiendo originar lesiones o enfermedades en las personas. Estos peligros se producen por prácticas de inocuidad alimentaria deficientes en cualquier punto de la cadena productiva. Un gran porcentaje de los daños que sufren las personas por el consumo de alimentos y bebidas provienen por los peligros físicos. Gran parte de las lesiones o enfermedades son daños en piezas dentales, molestias gastrointestinales, cortes y quemaduras en la boca y garganta. Los objetos extraños suelen ser: plásticos, piedras, metales, vidrios, cáscaras, maderas, y barros.'
    ws['E23'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['E23'].font = Font(size=6)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E23:S24']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A25:D26')
    ws['A25'] = 'Agentes Quimicos'
    ws['A25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A25'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A25:D26']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E25:S26')
    ws['E25'] = 'Los peligros químicos se generan de forma natural o se agregan durante cualquier etapa de procesamiento. Los agentes químicos son responsables de enfermedades crónicas, creación de alimentos tóxicos, presencia de sustancias venenosas en los alimentos, entre otros. Estos elementos dañinos ponen en alto riesgo la salud de las personas, por lo que es importante hacer énfasis en el sesion de la materia prima, además de conocer los procesos de producción, cosecha, procesamiento y almacenamiento.'
    ws['E25'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['E25'].font = Font(size=6)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E25:S26']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A27:D28')
    ws['A27'] = 'Agentes Biologicos'
    ws['A27'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A27'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A27:D28']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E27:S28')
    ws['E27'] = 'Entre los 3 tipos de peligros, el peligro biológico es el elemento que representa el mayor riesgo para los humanos. Los riesgos están compuestos por parásitos, bacterias, virus y hongos. Son agentes patógenos que se encuentran generalmente en el ambiente donde estos alimentos se producen. La mayoría de estos agentes son controlados mediante la cocción y una buena manipulación y almacenamiento.'
    ws['E27'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['E27'].font = Font(size=6)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E27:S28']:
        for cell in row:
            cell.border = border

    ws.merge_cells('C30:D30')
    ws['C30'] = 'Validado por:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['C30:D30']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E30:I30')
    ws['E30'] = mp_validado_por
    ws['E30'].alignment = Alignment(horizontal='center')
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E30:I30']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A32:B34')
    ws['A32'] = 'Cantidades Por el Largo:'
    ws['A32'].alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    ws['A32'].font = Font(size=10, bold=True)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A32:B34']:
        for cell in row:
            cell.border = border

    ws.merge_cells('C33:S33')
    ws['C33'] = cantidadesxlargoPP
    ws['C33'].alignment = Alignment(horizontal='center', vertical='center')

# ______________________________________________________________

    # Set the border style for the range C32:L32
    border_top = Border(top=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=32, max_row=32, min_col=3, max_col=19):
        for cell in row:
            cell.border = border_top

    border_right = Border(right=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=32, max_row=34, min_col=19, max_col=19):
        for cell in row:
            cell.border = border_right

    # Set the border style for the range C34:L34
    border_bottom = Border(bottom=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=34, max_row=34, min_col=3, max_col=19):
        for cell in row:
            cell.border = border_bottom

    cell_l32 = ws['S32']
    border_top_right = Border(top=Side(border_style='thin', color='000000'), right=Side(
        border_style='thin', color='000000'))
    cell_l32.border = border_top_right

    cell_l34 = ws['S34']
    border_bottom_right = Border(bottom=Side(
        border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
    cell_l34.border = border_bottom_right

    # Linea sepracion Prov - Lib
    border_left = Border(left=Side(border_style='thin', color='000000'))
    border_right = Border(left=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=8, max_col=8):
        for cell in row:
            cell.border = border_left

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=3, max_col=3):
        for cell in row:
            cell.border = border_right

    DiffT = PatternFill(start_color="D3D3D3",
                        end_color="D3D3D3", fill_type="solid")
    ws.merge_cells('I38:J38')
    ws['I38'] = 'Dif. M3:'
    ws['I38'].alignment = Alignment(horizontal='center')
    ws['I38'].font = Font(bold=True, size=10)
    ws['I38'].fill = DiffT

    ws.merge_cells('K38:L38')
    ws['K38'] = diferenciam3PP
    ws['K38'].alignment = Alignment(horizontal='center')
    ws['K38'].fill = DiffT

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=3, max_col=7):
        for cell in row:
            cell.border = border_bottom

    thick_border = Border(left=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    # Aplica el borde grueso a las celdas I38 a L38
    for row in ws.iter_rows(min_row=38, max_row=38, min_col=9, max_col=10):
        for cell in row:
            cell.border = thick_border

    thick_border = Border(right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))
    for row in ws.iter_rows(min_row=38, max_row=38, min_col=11, max_col=12):
        for cell in row:
            cell.border = thick_border

    DiffT = PatternFill(start_color="D3D3D3",
                        end_color="D3D3D3", fill_type="solid")
    ws.merge_cells('I40:J40')
    ws['I40'] = 'Dif. Pulgadas:'
    ws['I40'].alignment = Alignment(horizontal='center')
    ws['I40'].font = Font(bold=True, size=10)
    ws['I40'].fill = DiffT  # Aplica el relleno gris a la celda K33

    ws.merge_cells('K40:L40')
    ws['K40'] = diferenciapulgPP
    ws['K40'].alignment = Alignment(horizontal='center')
    ws['K40'].fill = DiffT

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=3, max_col=7):
        for cell in row:
            cell.border = border_bottom

    thick_border = Border(left=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    # Aplica el borde grueso a las celdas I38 a L38
    for row in ws.iter_rows(min_row=40, max_row=40, min_col=9, max_col=10):
        for cell in row:
            cell.border = thick_border

    thick_border = Border(right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))
    for row in ws.iter_rows(min_row=40, max_row=40, min_col=11, max_col=12):
        for cell in row:
            cell.border = thick_border

    DiffT = PatternFill(start_color="D3D3D3",
                        end_color="D3D3D3", fill_type="solid")
    ws.merge_cells('I42:J42')
    ws['I42'] = 'Dif. Unidades:'
    ws['I42'].alignment = Alignment(horizontal='center')
    ws['I42'].font = Font(bold=True, size=10)
    ws['I42'].fill = DiffT  # Aplica el relleno gris a la celda K33

    ws.merge_cells('K42:L42')
    ws['K42'] = diferenciacantidadPP
    ws['K42'].alignment = Alignment(horizontal='center')
    ws['K42'].fill = DiffT

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=3, max_col=7):
        for cell in row:
            cell.border = border_bottom

    thick_border = Border(left=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    # Aplica el borde grueso a las celdas I38 a L38
    for row in ws.iter_rows(min_row=42, max_row=42, min_col=9, max_col=10):
        for cell in row:
            cell.border = thick_border

    thick_border = Border(right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))
    for row in ws.iter_rows(min_row=42, max_row=42, min_col=11, max_col=12):
        for cell in row:
            cell.border = thick_border

    ws.merge_cells('C36:G36')
    ws['C36'] = proveedor
    ws['C36'].alignment = Alignment(horizontal='center')
    ws['C36'].font = Font(bold=True, size=11)
    border_bottom = Border(bottom=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=3, max_col=7):
        for cell in row:
            cell.border = border_bottom

    ws.merge_cells('C38:E38')
    ws['C38'] = 'M3 Total:'
    ws['C38'].alignment = Alignment(horizontal='center')
    ws['C38'].font = Font(size=10)
    ws['F38'] = m3PP

    ws.merge_cells('C40:E40')
    ws['C40'] = 'Pulgadas Total:'
    ws['C40'].alignment = Alignment(horizontal='center')
    ws['C40'].font = Font(size=10)
    ws['F40'] = pulgPP

    ws.merge_cells('C42:E42')
    ws['C42'] = 'Unidades Total:'
    ws['C42'].alignment = Alignment(horizontal='center')
    ws['C42'].font = Font(size=10)
    ws['F42'] = cantidadtotalPP

    ws.merge_cells('N36:R36')
    ws['N36'] = 'MADERAS INDUSTRIALES S.A.S'
    ws['N36'].alignment = Alignment(horizontal='center')
    ws['N36'].font = Font(bold=True, size=11)
    border_bottom = Border(bottom=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=14, max_col=18):
        for cell in row:
            cell.border = border_bottom

    ws.merge_cells('N38:P38')
    ws['N38'] = 'M3 Total:'
    ws['N38'].alignment = Alignment(horizontal='center')
    ws['N38'].font = Font(size=10)

    ws.merge_cells('Q38:R38')
    ws['Q38'] = m3_totalPP
    ws['Q38'].font = Font(size=10)

    ws.merge_cells('N40:P40')
    ws['N40'] = 'Pulgadas Total:'
    ws['N40'].alignment = Alignment(horizontal='center')
    ws['N40'].font = Font(size=10)

    ws.merge_cells('Q40:R40')
    ws['Q40'] = pulg_totalPP
    ws['Q40'].font = Font(size=10)

    ws.merge_cells('N42:P42')
    ws['N42'] = 'Unidades Total:'
    ws['N42'].alignment = Alignment(horizontal='center')
    ws['N42'].font = Font(size=10)

    ws.merge_cells('Q42:R42')
    ws['Q42'] = unidades_totalPP
    ws['Q42'].font = Font(size=10)

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=19, max_col=19):
        for cell in row:
            cell.border = border_left

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=14, max_col=14):
        for cell in row:
            cell.border = border_right

     # ******* TABLE PROVEEDOR ******

    # TABLE LINE 1
    ws.merge_cells('C45:G45')
    ws['C45'] = proveedor
    ws['C45'].alignment = Alignment(horizontal='center')
    ws['C45'].font = Font(bold=True, size=11)

    header_font = Font(size=8, bold=True)
    ws['C46'] = 'ID'
    ws['D46'] = 'Calibre'
    ws['E46'] = 'Ancho'
    ws['F46'] = 'Largo'
    ws['G46'] = 'Unidades'

    # TABLE LINE 1
    ws['C46'].font = header_font
    ws['D46'].font = header_font
    ws['E46'].font = header_font
    ws['F46'].font = header_font
    ws['G46'].font = header_font

    # Split the diametroTROZA and largoTROZA strings into lists
    calibrePP = calibrePP.split(',')
    anchoPPArray = anchoPP.split(',')
    largoPPArray = largoPP.split(',')
    cantidadPPArray = cantidadPP.split(',')

    max_rows_per_cell = 40  # Maximum rows per cell
    current_row = 47  # Starting row for data
    data_counter = 1

    for i, (calibre, ancho, largo, unds) in enumerate(zip(calibrePP, anchoPPArray, largoPPArray, cantidadPPArray), start=1):
        # Remove any leading/trailing spaces and convert to float
        calibre = float(calibre.strip())
        ancho = float(ancho.strip())
        largo = float(largo.strip())
        unds = float(unds.strip())
        if i % max_rows_per_cell == 1 and i > max_rows_per_cell:
            current_row = 47  # Reset to the starting row of the next cell

        # TABLES LINE 1
        else:
            ws['C{}'.format(current_row)] = data_counter
            ws['D{}'.format(current_row)] = calibre
            ws['E{}'.format(current_row)] = ancho
            ws['F{}'.format(current_row)] = largo
            ws['G{}'.format(current_row)] = unds

            data_counter += 1
            current_row += 1

    # Apply styling
    border = Border(left=Side(style='medium', color='000000'),  # Medium weight and black color
                    # Medium weight and black color
                    right=Side(style='medium', color='000000'),
                    # Thin weight and black color
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))  # Thin weight and black color

    # TABLES LINE 1
    for col_num in range(3, 8):
        for row_num in range(46, 61 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # ******* TABLE DIFF ******

    # TABLE LINE 1
    ws.merge_cells('J46:K46')
    ws['J46'] = "Diff. Unds"
    ws['J46'].alignment = Alignment(horizontal='center')
    ws['J46'].font = Font(bold=True, size=11)

    header_font = Font(size=8, bold=True)
    ws['K47'] = 'Unds'

    # TABLE LINE 1
    ws['J46'].font = header_font
    ws['K46'].font = header_font

    # Split the diametroTROZA and largoTROZA strings into lists
    DiffUndsArray = diff_unds.split(',')

    max_rows_per_cell = 40  # Maximum rows per cell
    current_row = 47  # Starting row for data
    data_counter = 1

    for i, Diffunds in enumerate(DiffUndsArray, start=1):
        if i % max_rows_per_cell == 1 and i > max_rows_per_cell:
            current_row = 47  # Reiniciar a la fila de inicio de la siguiente celda

        # Convierte Diffunds a un número entero (si es un número)
        try:
            Diffunds = int(Diffunds)
        except ValueError:
            # En caso de que no se pueda convertir a entero, maneja el error o elige un valor predeterminado
            Diffunds = 0  # Por ejemplo, asigna 0 si no se puede convertir a entero

        ws['J{}'.format(current_row)] = Diffunds

        data_counter += 1
        current_row += 1

    # Apply styling
    border = Border(left=Side(style='medium', color='000000'),  # Medium weight and black color
                    # Medium weight and black color
                    right=Side(style='medium', color='000000'),
                    # Thin weight and black color
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))  # Thin weight and black color

    # TABLES LINE 1
    for col_num in range(10, 12):
        for row_num in range(46, 61 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for row_num in range(47, 61 + 1):
        # Itera a través de las columnas J y K
        for col_num in range(10, 12):
            col_letter = get_column_letter(col_num)
            cell = ws['{}{}'.format(col_letter, row_num)]
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Combina las celdas en pares (J47-K47, J48-K48, etc.) hasta la fila 61
    for row_num in range(47, 61 + 1):
        col_letter1 = get_column_letter(10)  # Columna J
        col_letter2 = get_column_letter(11)  # Columna K
        ws.merge_cells('{}{}:{}{}'.format(
            col_letter1, row_num, col_letter2, row_num))

    # ******* TABLE LIBERADO ******

    # TABLE LINE 1
    ws.merge_cells('N45:R45')
    ws['N45'] = "MADERAS INDSUTRIALES S.A.S"
    ws['N45'].alignment = Alignment(horizontal='center')
    ws['N45'].font = Font(bold=True, size=11)

    header_font = Font(size=8, bold=True)
    ws['N46'] = 'ID'
    ws['O46'] = 'Calibre'
    ws['P46'] = 'Ancho'
    ws['Q46'] = 'Largo'
    ws['R46'] = 'Unidades'

    # TABLE LINE 1
    ws['N46'].font = header_font
    ws['O46'].font = header_font
    ws['P46'].font = header_font
    ws['Q46'].font = header_font
    ws['R46'].font = header_font

    dataArray = cantidadesxlargoPP.split('/')

    max_rows_per_cell = 40  # Maximum rows per cell
    current_row = 47  # Starting row for data
    data_counter = 1

    for i, rowData in enumerate(dataArray, start=1):
        if i % max_rows_per_cell == 1 and i > max_rows_per_cell:
            current_row = 47  # Reset to the starting row of the next cell

        # Split rowData by '/' to get individual items
        items = rowData.split('/')

        for item in items:
            # Split item by 'x' to get calibre, ancho, largo, unds
            parts = item.split('x')
            calibre = parts[0].strip()
            ancho = parts[1].strip()
            largo_unds = parts[2].strip()

            # Split largo_unds by ':' to get largo and unds
            largo, unds = largo_unds.split(':')
            largo = largo.strip()
            unds = unds.strip()

            # TABLES LINE 1
            ws['N{}'.format(current_row)] = data_counter
            ws['O{}'.format(current_row)] = calibre
            ws['P{}'.format(current_row)] = ancho
            ws['Q{}'.format(current_row)] = largo
            ws['R{}'.format(current_row)] = unds

            data_counter += 1
            current_row += 1

    # Apply styling
    border = Border(left=Side(style='medium', color='000000'),  # Medium weight and black color
                    # Medium weight and black color
                    right=Side(style='medium', color='000000'),
                    # Thin weight and black color
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))  # Thin weight and black color

    # TABLES LINE 1
    for col_num in range(14, 19):
        for row_num in range(46, 61 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border


# ______________________________________________________________
    # Add the image to cell D1
    img = Image('src/static/img/splash.png')
    img.width = 80
    img.height = 80
    img.left = 40  # Agregar un margen izquierdo de 40 unidades
    ws.add_image(img, 'A1')
    ws.merge_cells('A1:B4')

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A1:B4']:
        for cell in row:
            cell.border = border

    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)

    # Set the response headers to send the Excel file to the client
    response = make_response(virtual_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=LOTE_PP_{}.xlsx'.format(
        lote)

    return response

# **************************************************************************************************************************
# *************************************************************BLOQUE VAL EXCEL*************************************************************
# **************************************************************************************************************************


@app.route('/generate_excelBLOQUEVal', methods=['POST'])
@login_required
def generate_excelBLOQUEVal():
    # Retrieve form data from the request
    lote = request.form['lote']
    fecha_ingreso = request.form['fecha_ingreso']
    fecha = request.form['fecha']
    proveedor = request.form['proveedor']
    mp_gorgojo = request.form['mp_gorgojo']
    mp_partida = request.form['mp_partida']
    mp_fisicos = request.form['mp_fisicos']
    mp_quimicos = request.form['mp_quimicos']
    mp_biologico = request.form['mp_biologico']
    cantidadesxlargoBLOQUE = request.form['cantidadesxlargoBLOQUE']
    mp_validado_por = request.form['mp_validado_por']

    diferenciapulgBLOQUE = request.form['diferenciapulgBLOQUE']

    pulgadasBLOQUE = request.form['pulgadasBLOQUE']

    pulg_totalBLOQUE = request.form['pulg_totalBLOQUE']
    m3_totalBLOQUE = request.form['m3_totalBLOQUE']
    unidades_totalBLOQUE = request.form['unidades_totalBLOQUE']

    lado1BLOQUE = request.form['lado1BLOQUE']
    lado2BLOQUE = request.form['lado2BLOQUE']
    lengthBLOQUE = request.form['lengthBLOQUE']

    # Add more fields as needed

    # Create a new workbook
    wb = Workbook()
    ws = wb.active

    # LINEAS DE CUADRICULA DESACTIVADO
    sheet = wb.active
    sheet.sheet_view.showGridLines = False

    ancho_id = 4
    columna = ws.column_dimensions['A']
    columna.width = ancho_id
    columna = ws.column_dimensions['F']
    columna.width = ancho_id
    columna = ws.column_dimensions['K']
    columna.width = ancho_id
    columna = ws.column_dimensions['P']
    columna.width = ancho_id

    ancho_largo = 9
    columna = ws.column_dimensions['D']
    columna.width = ancho_largo
    columna = ws.column_dimensions['I']
    columna.width = ancho_largo
    columna = ws.column_dimensions['N']
    columna.width = ancho_largo
    columna = ws.column_dimensions['S']
    columna.width = ancho_largo

    # Write the form data to the worksheet
    ws.merge_cells('C1:Q4')  # Merge cells E1 to J2
    ws['C1'] = 'FORMATO INFORME LIBERADO BLOQUE'

    # Set alignment and formatting for the merged cell
    merged_cell = ws['C1']
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = Font(bold=True, size=14)

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['C1:Q4']:
        for cell in row:
            cell.border = border

    ws['R1'] = 'Codigo'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R1'].border = border

    ws['S1'] = 'FSAA-VDT-01'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S1'].border = border

    ws['R2'] = 'Vigencia'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R2'].border = border

    ws['S2'] = '21/06/2022'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S2'].border = border

    ws['R3'] = 'Versión'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R3'].border = border

    ws['S3'] = '2'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S3'].border = border

    ws['R4'] = 'Pag'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R4'].border = border

    ws['S4'] = '1 de 1'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S4'].border = border

    font_size = 8
    cell_range = ['S1', 'R1', 'S2', 'R2', 'S3', 'R3', 'S4', 'R4']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)

    ws['C5'] = 'LOTE:'
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    for col in range(3, 5):  # Columnas C y D
        for row in range(5, 6):  # Fila 5
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = Font(size=13)

    # Merge y estilo de la celda E5
    ws.merge_cells('D5:E5')
    ws['D5'] = lote
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    for col in range(5, 6):  # Columnas E y F
        for row in range(5, 6):  # Fila 5
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = Font(size=13)

    # Set alignment for the cell containing the lote value
    ws['B1'].alignment = Alignment(horizontal='left')

    # Write other form data to the worksheet
    ws.merge_cells('L7:M7')
    ws['L7'] = 'Fecha Ingreso:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['L7:M7']:
        for cell in row:
            cell.border = border

    ws.merge_cells('N7:O7')
    ws['N7'] = fecha_ingreso
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    ws['N7'].border = border
    for row in ws['N7:O7']:
        for cell in row:
            cell.border = border

    ws.merge_cells('L8:M8')
    ws['L8'] = 'Fecha Validación:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['L8:M8']:
        for cell in row:
            cell.border = border

    ws.merge_cells('N8:O8')
    ws['N8'] = fecha
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    ws['N8'].border = border
    for row in ws['N8:O8']:
        for cell in row:
            cell.border = border

    ws.merge_cells('D7:E7')
    ws['D7'] = 'Proveedor:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['D7:F7']:
        for cell in row:
            cell.border = border

    ws.merge_cells('F7:I7')
    ws['F7'] = proveedor
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['F7:I7']:
        for cell in row:
            cell.border = border


# *************** CRITERIO DE ACEPTACION Y RECHAZO ******************
    ws.merge_cells('A11:F11')
    ws['A11'] = 'CRITERIO DE ACEPTACION Y RECHAZO'
    ws['A11'].alignment = Alignment(horizontal='center')
    ws['A11'].font = Font(bold=True, size=8)

    # Crear el estilo de borde
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A11:F11']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A12:F12')
    ws['A12'] = 'La madera cumple con las medidas vs remisión?'
    # Crear el estilo de borde
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A12:F12']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A13:F13')
    ws['A13'] = 'La madera cumple con las cantidades vs remisión?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A13:F13']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A14:F14')
    ws['A14'] = 'La madera presenta gorgojo?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A14:F14']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A15:F15')
    ws['A15'] = 'La madera esta partida o quebrada?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A15:F15']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A16:F16')
    ws['A16'] = 'La madera tiene agentes contaminantes físicos?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A16:F16']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A17:F17')
    ws['A17'] = 'La madera tiene agentes contaminantes químicos?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A17:F17']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A18:F18')
    ws['A18'] = 'La madera tiene agentes contaminantes biológicos?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A18:F18']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A19:F19')
    ws['A19'] = 'La madera tiene salvoconducto de movilización?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A19:F19']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A20:F20')
    ws['A20'] = 'Requiere informe no conformidad para el proveedor?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A20:F20']:
        for cell in row:
            cell.border = border

    font_size = 8
    cell_range = ['A12', 'A13', 'A14', 'A15',
                  'A16', 'A17', 'A18', 'A19', 'A20']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)

    ws.merge_cells('G11:L11')
    ws['G11'] = 'Marque con una (X)'
    ws['G11'].alignment = Alignment(horizontal='center')
    ws['G11'].font = Font(bold=True, size=8)

    # Crear el estilo de borde
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['G11:L11']:
        for cell in row:
            cell.border = border

    ws['G12'] = 'Si'
    ws['G12'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='808080'))
    ws['G12'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H12'].border = border
    ws.merge_cells('H12:I12')

    ws['G13'] = 'Si'
    ws['G13'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G13'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H13'].border = border
    ws.merge_cells('H13:I13')

    ws['G14'] = 'Si'
    ws['G14'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G14'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H14'].border = border
    ws.merge_cells('H14:I14')

    ws['G15'] = 'Si'
    ws['G15'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G15'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H15'].border = border
    ws.merge_cells('H15:I15')

    ws['G16'] = 'Si'
    ws['G16'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G16'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H16'].border = border
    ws.merge_cells('H16:I16')

    ws['G17'] = 'Si'
    ws['G17'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G17'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H17'].border = border
    ws.merge_cells('H17:I17')

    ws['G18'] = 'Si'
    ws['G18'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G18'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H12'].border = border
    ws.merge_cells('H18:I18')

    ws['G19'] = 'Si'
    ws['G19'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G19'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H19'].border = border
    ws.merge_cells('H19:I19')

    ws['G20'] = 'Si'
    ws['G20'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='000000'))
    ws['G20'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='000000'))

    ws['H20'].border = border
    ws.merge_cells('H20:I20')

    # _____________NO______________
    ws['J12'] = 'No'
    ws['J12'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J12'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K12'].border = border
    ws.merge_cells('K12:L12')

    ws['J13'] = 'No'
    ws['J13'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J13'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K13'].border = border
    ws.merge_cells('K13:L13')

    ws['J14'] = 'No'
    ws['J14'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J14'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K14'].border = border
    ws.merge_cells('K14:L14')

    ws['J15'] = 'No'
    ws['J15'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J15'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K15'].border = border
    ws.merge_cells('K15:L15')

    ws['J16'] = 'No'
    ws['J16'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J16'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K16'].border = border
    ws.merge_cells('K16:L16')

    ws['J17'] = 'No'
    ws['J17'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J17'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K17'].border = border
    ws.merge_cells('K17:L17')

    ws['J18'] = 'No'
    ws['J18'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J18'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K18'].border = border
    ws.merge_cells('K18:L18')

    ws['J19'] = 'No'
    ws['J19'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J19'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K19'].border = border
    ws.merge_cells('K19:L19')

    ws['J20'] = 'No'
    ws['J20'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='000000'))
    ws['J20'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='000000'))

    ws['K20'].border = border
    ws.merge_cells('K20:L20')

    font_size = 8
    cell_range = ['G12', 'G13', 'G14', 'G15', 'G16', 'G17', 'G18', 'G19',
                  'G20', 'J12', 'J13', 'J14', 'J15', 'J16', 'J17', 'J18', 'J19', 'J20']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)

# OBSERVACIONES
    ws.merge_cells('M11:S11')
    ws['M11'] = 'OBSERVACIONES'
    ws['M11'].alignment = Alignment(horizontal='center')
    ws['M11'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in ws['M11:S11']:
        for cell in row:
            cell.border = border

    # _____COMETARIOS______

    ws.merge_cells('M12:S12')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M12:S12']:
        for cell in row:
            cell.border = border

    ws.merge_cells('M13:S13')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M13:S13']:
        for cell in row:
            cell.border = border

    ws['M14'] = mp_gorgojo
    ws.merge_cells('M14:S14')

    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M14:S14']:
        for cell in row:
            cell.border = border

    ws['M15'] = mp_partida
    ws.merge_cells('M15:S15')

    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M15:S15']:
        for cell in row:
            cell.border = border

    ws['M16'] = mp_fisicos
    ws.merge_cells('M16:S16')

    ws.merge_cells('M16:S16')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M16:S16']:
        for cell in row:
            cell.border = border

    ws['M17'] = mp_quimicos
    ws.merge_cells('M17:S17')

    ws.merge_cells('M17:S17')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M17:S17']:
        for cell in row:
            cell.border = border

    ws['M18'] = mp_biologico
    ws.merge_cells('M18:S18')

    ws.merge_cells('M18:S18')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M18:S18']:
        for cell in row:
            cell.border = border

    ws.merge_cells('M19:S19')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M19:S19']:
        for cell in row:
            cell.border = border

    ws.merge_cells('M20:S20')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in ws['M20:S20']:
        for cell in row:
            cell.border = border

    font_size = 8
    cell_range = ['M14', 'M15', 'M16', 'M17', 'M18']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)
# ______________________________________________________________

    ws.merge_cells('A22:S22')
    ws['A22'] = 'PONDERACION'
    ws['A22'].alignment = Alignment(horizontal='center')
    ws['A22'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in ws['A22:S22']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A23:D24')
    ws['A23'] = 'Agentes Fisícos'
    ws['A23'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A23'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A23:D24']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E23:S24')
    ws['E23'] = 'Los peligros físicos constan de objetos extraños presentes en los alimentos, pudiendo originar lesiones o enfermedades en las personas. Estos peligros se producen por prácticas de inocuidad alimentaria deficientes en cualquier punto de la cadena productiva. Un gran porcentaje de los daños que sufren las personas por el consumo de alimentos y bebidas provienen por los peligros físicos. Gran parte de las lesiones o enfermedades son daños en piezas dentales, molestias gastrointestinales, cortes y quemaduras en la boca y garganta. Los objetos extraños suelen ser: plásticos, piedras, metales, vidrios, cáscaras, maderas, y barros.'
    ws['E23'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['E23'].font = Font(size=6)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E23:S24']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A25:D26')
    ws['A25'] = 'Agentes Quimicos'
    ws['A25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A25'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A25:D26']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E25:S26')
    ws['E25'] = 'Los peligros químicos se generan de forma natural o se agregan durante cualquier etapa de procesamiento. Los agentes químicos son responsables de enfermedades crónicas, creación de alimentos tóxicos, presencia de sustancias venenosas en los alimentos, entre otros. Estos elementos dañinos ponen en alto riesgo la salud de las personas, por lo que es importante hacer énfasis en el sesion de la materia prima, además de conocer los procesos de producción, cosecha, procesamiento y almacenamiento.'
    ws['E25'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['E25'].font = Font(size=6)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E25:S26']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A27:D28')
    ws['A27'] = 'Agentes Biologicos'
    ws['A27'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A27'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A27:D28']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E27:S28')
    ws['E27'] = 'Entre los 3 tipos de peligros, el peligro biológico es el elemento que representa el mayor riesgo para los humanos. Los riesgos están compuestos por parásitos, bacterias, virus y hongos. Son agentes patógenos que se encuentran generalmente en el ambiente donde estos alimentos se producen. La mayoría de estos agentes son controlados mediante la cocción y una buena manipulación y almacenamiento.'
    ws['E27'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['E27'].font = Font(size=6)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E27:S28']:
        for cell in row:
            cell.border = border

    ws.merge_cells('D30:E30')
    ws['D30'] = 'Validado por:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['D30:E30']:
        for cell in row:
            cell.border = border

    ws.merge_cells('F30:H30')
    ws['F30'] = mp_validado_por
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['F30:H30']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A32:B34')
    ws['A32'] = 'Cantidades Por el Largo:'
    ws['A32'].alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    ws['A32'].font = Font(size=10, bold=True)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A32:B34']:
        for cell in row:
            cell.border = border

    ws.merge_cells('C33:S33')
    ws['C33'] = cantidadesxlargoBLOQUE
    ws['C33'].alignment = Alignment(horizontal='center', vertical='center')

# ______________________________________________________________

    # Set the border style for the range C32:L32
    border_top = Border(top=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=32, max_row=32, min_col=3, max_col=19):
        for cell in row:
            cell.border = border_top

    border_right = Border(right=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=32, max_row=34, min_col=19, max_col=19):
        for cell in row:
            cell.border = border_right

    # Set the border style for the range C34:L34
    border_bottom = Border(bottom=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=34, max_row=34, min_col=3, max_col=19):
        for cell in row:
            cell.border = border_bottom

    cell_l32 = ws['S32']
    border_top_right = Border(top=Side(border_style='thin', color='000000'), right=Side(
        border_style='thin', color='000000'))
    cell_l32.border = border_top_right

    cell_l34 = ws['S34']
    border_bottom_right = Border(bottom=Side(
        border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
    cell_l34.border = border_bottom_right

    # Linea sepracion Prov - Lib
    border_left = Border(left=Side(border_style='thin', color='000000'))
    border_right = Border(left=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=8, max_col=8):
        for cell in row:
            cell.border = border_left

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=3, max_col=3):
        for cell in row:
            cell.border = border_right

    DiffB = PatternFill(start_color="D3D3D3",
                        end_color="D3D3D3", fill_type="solid")
    ws.merge_cells('I38:J38')
    ws['I38'] = 'Dif. Pulgadas:'
    ws['I38'].alignment = Alignment(horizontal='center')
    ws['I38'].font = Font(bold=True, size=10)
    ws['I38'].fill = DiffB

    ws.merge_cells('K38:L38')
    ws['K38'] = diferenciapulgBLOQUE
    ws['K38'].fill = DiffB

    ws.merge_cells('C36:G36')
    ws['C36'] = proveedor
    ws['C36'].alignment = Alignment(horizontal='center')
    ws['C36'].font = Font(bold=True, size=11)
    border_bottom = Border(bottom=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=3, max_col=7):
        for cell in row:
            cell.border = border_bottom

    thick_border = Border(left=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    # Aplica el borde grueso a las celdas I38 a L38
    for row in ws.iter_rows(min_row=38, max_row=38, min_col=9, max_col=10):
        for cell in row:
            cell.border = thick_border

    thick_border = Border(right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))
    for row in ws.iter_rows(min_row=38, max_row=38, min_col=11, max_col=12):
        for cell in row:
            cell.border = thick_border

    ws.merge_cells('D38:E38')
    ws['D38'] = 'Pulgadas Total:'
    ws['D38'].alignment = Alignment(horizontal='center')
    ws['D38'].font = Font(size=10)
    ws['F38'] = pulgadasBLOQUE

    ws.merge_cells('N36:R36')
    ws['N36'] = 'MADERAS INDSUTRIALES S.A.S'
    ws['N36'].alignment = Alignment(horizontal='center')
    ws['N36'].font = Font(bold=True, size=11)
    border_bottom = Border(bottom=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=14, max_col=18):
        for cell in row:
            cell.border = border_bottom

    ws.merge_cells('N38:O38')
    ws['N38'] = 'Pulgadas Total:'
    ws['N38'].alignment = Alignment(horizontal='center')
    ws['N38'].font = Font(size=10)

    ws.merge_cells('P38:R38')
    ws['P38'] = pulg_totalBLOQUE
    ws['P38'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('N40:O40')
    ws['N40'] = 'M3 Total:'
    ws['N40'].alignment = Alignment(horizontal='center')
    ws['N40'].font = Font(size=10)

    ws.merge_cells('P40:R40')
    ws['P40'] = m3_totalBLOQUE
    ws['P40'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('N42:O42')
    ws['N42'] = 'Unidades Total:'
    ws['N42'].alignment = Alignment(horizontal='center')
    ws['N42'].font = Font(size=10)

    ws.merge_cells('P42:R42')
    ws['P42'] = unidades_totalBLOQUE
    ws['P42'].alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=19, max_col=19):
        for cell in row:
            cell.border = border_left

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=14, max_col=14):
        for cell in row:
            cell.border = border_right


# ______________________________________________________________
    # Add the image to cell D1
    img = Image('src/static/img/splash.png')
    img.width = 80
    img.height = 80
    img.left = 40  # Agregar un margen izquierdo de 40 unidades
    ws.add_image(img, 'A1')
    ws.merge_cells('A1:B4')

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A1:B4']:
        for cell in row:
            cell.border = border

    # ******* TABLE LIBERADO ******

    # TABLE LINE 1
    ws.merge_cells('A45:I45')
    ws['A45'] = "MADERAS INDSUTRIALES S.A.S"
    ws['A45'].alignment = Alignment(horizontal='center')
    ws['A45'].font = Font(bold=True, size=11)

    ws.merge_cells('K45:S45')
    ws['K45'] = "MADERAS INDSUTRIALES S.A.S"
    ws['K45'].alignment = Alignment(horizontal='center')
    ws['K45'].font = Font(bold=True, size=11)

    # TABLE LINE 2
    ws.merge_cells('A88:I88')
    ws['A88'] = "MADERAS INDSUTRIALES S.A.S"
    ws['A88'].alignment = Alignment(horizontal='center')
    ws['A88'].font = Font(bold=True, size=11)

    ws.merge_cells('K88:S88')
    ws['K88'] = "MADERAS INDSUTRIALES S.A.S"
    ws['K88'].alignment = Alignment(horizontal='center')
    ws['K88'].font = Font(bold=True, size=11)

    # TABLE LINE 3
    ws.merge_cells('A131:I131')
    ws['A131'] = "MADERAS INDSUTRIALES S.A.S"
    ws['A131'].alignment = Alignment(horizontal='center')
    ws['A131'].font = Font(bold=True, size=11)

    ws.merge_cells('K131:S131')
    ws['K131'] = "MADERAS INDSUTRIALES S.A.S"
    ws['K131'].alignment = Alignment(horizontal='center')
    ws['K131'].font = Font(bold=True, size=11)

    # TABLE LINE 4
    ws.merge_cells('A174:I174')
    ws['A174'] = "MADERAS INDSUTRIALES S.A.S"
    ws['A174'].alignment = Alignment(horizontal='center')
    ws['A174'].font = Font(bold=True, size=11)

    ws.merge_cells('K174:S174')
    ws['K174'] = "MADERAS INDSUTRIALES S.A.S"
    ws['K174'].alignment = Alignment(horizontal='center')
    ws['K174'].font = Font(bold=True, size=11)

    # TABLE LINE 5
    ws.merge_cells('A217:I217')
    ws['A217'] = "MADERAS INDSUTRIALES S.A.S"
    ws['A217'].alignment = Alignment(horizontal='center')
    ws['A217'].font = Font(bold=True, size=11)

    ws.merge_cells('K217:S217')
    ws['K217'] = "MADERAS INDSUTRIALES S.A.S"
    ws['K217'].alignment = Alignment(horizontal='center')
    ws['K217'].font = Font(bold=True, size=11)

    # TABLE LINE 6
    ws.merge_cells('A260:I260')
    ws['A260'] = "MADERAS INDSUTRIALES S.A.S"
    ws['A260'].alignment = Alignment(horizontal='center')
    ws['A260'].font = Font(bold=True, size=11)

    ws.merge_cells('K260:S260')
    ws['K260'] = "MADERAS INDSUTRIALES S.A.S"
    ws['K260'].alignment = Alignment(horizontal='center')
    ws['K260'].font = Font(bold=True, size=11)

    header_font = Font(size=8, bold=True)
    ws['A46'] = 'ID'
    ws['B46'] = 'Lado 1'
    ws['C46'] = 'Lado 2'
    ws['D46'] = 'Largo'

    # TABLE LINE 1
    ws['A46'].font = header_font
    ws['B46'].font = header_font
    ws['C46'].font = header_font
    ws['D46'].font = header_font

    ws['F46'].font = header_font
    ws['G46'].font = header_font
    ws['H46'].font = header_font
    ws['I46'].font = header_font

    ws['K46'].font = header_font
    ws['L46'].font = header_font
    ws['M46'].font = header_font
    ws['N46'].font = header_font

    ws['P46'].font = header_font
    ws['Q46'].font = header_font
    ws['R46'].font = header_font
    ws['S46'].font = header_font

    # TABLE LINE 2
    ws['A89'].font = header_font
    ws['B89'].font = header_font
    ws['C89'].font = header_font
    ws['D89'].font = header_font

    ws['F89'].font = header_font
    ws['G89'].font = header_font
    ws['H89'].font = header_font
    ws['I89'].font = header_font

    ws['K89'].font = header_font
    ws['L89'].font = header_font
    ws['M89'].font = header_font
    ws['N89'].font = header_font

    ws['P89'].font = header_font
    ws['Q89'].font = header_font
    ws['R89'].font = header_font
    ws['S89'].font = header_font

    # TABLE LINE 3
    ws['A132'].font = header_font
    ws['B132'].font = header_font
    ws['C132'].font = header_font
    ws['D132'].font = header_font

    ws['F132'].font = header_font
    ws['G132'].font = header_font
    ws['H132'].font = header_font
    ws['I132'].font = header_font

    ws['K132'].font = header_font
    ws['L132'].font = header_font
    ws['M132'].font = header_font
    ws['N132'].font = header_font

    ws['P132'].font = header_font
    ws['Q132'].font = header_font
    ws['R132'].font = header_font
    ws['S132'].font = header_font

    # TABLE LINE 4
    ws['A175'].font = header_font
    ws['B175'].font = header_font
    ws['C175'].font = header_font
    ws['D175'].font = header_font

    ws['F175'].font = header_font
    ws['G175'].font = header_font
    ws['H175'].font = header_font
    ws['I175'].font = header_font

    ws['K175'].font = header_font
    ws['L175'].font = header_font
    ws['M175'].font = header_font
    ws['N175'].font = header_font

    ws['P175'].font = header_font
    ws['Q175'].font = header_font
    ws['R175'].font = header_font
    ws['S175'].font = header_font

    # TABLE LINE 5
    ws['A218'].font = header_font
    ws['B218'].font = header_font
    ws['C218'].font = header_font
    ws['D218'].font = header_font

    ws['F218'].font = header_font
    ws['G218'].font = header_font
    ws['H218'].font = header_font
    ws['I218'].font = header_font

    ws['K218'].font = header_font
    ws['L218'].font = header_font
    ws['M218'].font = header_font
    ws['N218'].font = header_font

    ws['P218'].font = header_font
    ws['Q218'].font = header_font
    ws['R218'].font = header_font
    ws['S218'].font = header_font

    # TABLE LINE 6
    ws['A261'].font = header_font
    ws['B261'].font = header_font
    ws['C261'].font = header_font
    ws['D261'].font = header_font

    ws['F261'].font = header_font
    ws['G261'].font = header_font
    ws['H261'].font = header_font
    ws['I261'].font = header_font

    ws['K261'].font = header_font
    ws['L261'].font = header_font
    ws['M261'].font = header_font
    ws['N261'].font = header_font

    ws['P261'].font = header_font
    ws['Q261'].font = header_font
    ws['R261'].font = header_font
    ws['S261'].font = header_font

    lado1BLOQUEArray = lado1BLOQUE.split(',')
    lado2BLOQUEArray = lado2BLOQUE.split(',')
    lengthBLOQUEArray = lengthBLOQUE.split(',')

    max_rows_per_cell = 40  # Maximum rows per cell
    current_row = 47  # Starting row for data
    data_counter = 1

    for i, (lado1, lado2, length) in enumerate(zip(lado1BLOQUEArray, lado2BLOQUEArray, lengthBLOQUEArray), start=1):
        # Remove any leading/trailing spaces and convert to float
        lado1 = float(lado1.strip())
        lado2 = float(lado2.strip())
        length = float(length.strip())
        if i % max_rows_per_cell == 1 and i > max_rows_per_cell:
            current_row = 47  # Reset to the starting row of the next cell
            ws['F46'] = 'ID'
            ws['G46'] = 'Lado 1'
            ws['H46'] = 'Lado 2'
            ws['I46'] = 'Largo'
            ws['K46'] = 'ID'
            ws['L46'] = 'Lado 1'
            ws['M46'] = 'Lado 2'
            ws['N46'] = 'Largo'
            ws['P46'] = 'ID'
            ws['Q46'] = 'Lado 1'
            ws['R46'] = 'Lado 2'
            ws['S46'] = 'Largo'

            ws['A89'] = 'ID'
            ws['B89'] = 'Lado 1'
            ws['C89'] = 'Lado 2'
            ws['D89'] = 'Largo'
            ws['F89'] = 'ID'
            ws['G89'] = 'Lado 1'
            ws['H89'] = 'Lado 2'
            ws['I89'] = 'Largo'
            ws['K89'] = 'ID'
            ws['L89'] = 'Lado 1'
            ws['M89'] = 'Lado 2'
            ws['N89'] = 'Largo'
            ws['P89'] = 'ID'
            ws['Q89'] = 'Lado 1'
            ws['R89'] = 'Lado 2'
            ws['S89'] = 'Largo'

            ws['A132'] = 'ID'
            ws['B132'] = 'Lado 1'
            ws['C132'] = 'Lado 2'
            ws['D132'] = 'Largo'
            ws['F132'] = 'ID'
            ws['G132'] = 'Lado 1'
            ws['H132'] = 'Lado 2'
            ws['I132'] = 'Largo'
            ws['K132'] = 'ID'
            ws['L132'] = 'Lado 1'
            ws['M132'] = 'Lado 2'
            ws['N132'] = 'Largo'
            ws['P132'] = 'ID'
            ws['Q132'] = 'Lado 1'
            ws['R132'] = 'Lado 2'
            ws['S132'] = 'Largo'

            ws['A175'] = 'ID'
            ws['B175'] = 'Lado 1'
            ws['C175'] = 'Lado 2'
            ws['D175'] = 'Largo'
            ws['F175'] = 'ID'
            ws['G175'] = 'Lado 1'
            ws['H175'] = 'Lado 2'
            ws['I175'] = 'Largo'
            ws['K175'] = 'ID'
            ws['L175'] = 'Lado 1'
            ws['M175'] = 'Lado 2'
            ws['N175'] = 'Largo'
            ws['P175'] = 'ID'
            ws['Q175'] = 'Lado 1'
            ws['R175'] = 'Lado 2'
            ws['S175'] = 'Largo'

            ws['A218'] = 'ID'
            ws['B218'] = 'Lado 1'
            ws['C218'] = 'Lado 2'
            ws['D218'] = 'Largo'
            ws['F218'] = 'ID'
            ws['G218'] = 'Lado 1'
            ws['H218'] = 'Lado 2'
            ws['I218'] = 'Largo'
            ws['K218'] = 'ID'
            ws['L218'] = 'Lado 1'
            ws['M218'] = 'Lado 2'
            ws['N218'] = 'Largo'
            ws['P218'] = 'ID'
            ws['Q218'] = 'Lado 1'
            ws['R218'] = 'Lado 2'
            ws['S218'] = 'Largo'

            ws['A261'] = 'ID'
            ws['B261'] = 'Lado 1'
            ws['C261'] = 'Lado 2'
            ws['D261'] = 'Largo'
            ws['F261'] = 'ID'
            ws['G261'] = 'Lado 1'
            ws['H261'] = 'Lado 2'
            ws['I261'] = 'Largo'
            ws['K261'] = 'ID'
            ws['L261'] = 'Lado 1'
            ws['M261'] = 'Lado 2'
            ws['N261'] = 'Largo'
            ws['P261'] = 'ID'
            ws['Q261'] = 'Lado 1'
            ws['R261'] = 'Lado 2'
            ws['S261'] = 'Largo'

# TABLES LINE 1
        if i > max_rows_per_cell and i <= 80:
            ws['F{}'.format(current_row)] = i
            ws['G{}'.format(current_row)] = lado1
            ws['H{}'.format(current_row)] = lado2
            ws['I{}'.format(current_row)] = length
            current_row += 1
        elif i > 80 and i <= 120:
            ws['K{}'.format(current_row)] = i
            ws['L{}'.format(current_row)] = lado1
            ws['M{}'.format(current_row)] = lado2
            ws['N{}'.format(current_row)] = length
            current_row += 1
        elif i > 120 and i <= 160:
            ws['P{}'.format(47 + i - 121)] = i
            ws['Q{}'.format(47 + i - 121)] = lado1
            ws['R{}'.format(47 + i - 121)] = lado2
            ws['S{}'.format(47 + i - 121)] = length
            current_row += 1

# TABLES LINE 2
        elif i > 160 and i <= 200:
            ws['A{}'.format(90 + i - 161)] = i
            ws['B{}'.format(90 + i - 161)] = lado1
            ws['C{}'.format(90 + i - 161)] = lado2
            ws['D{}'.format(90 + i - 161)] = length
            current_row += 1
        elif i > 200 and i <= 240:
            ws['F{}'.format(90 + i - 201)] = i
            ws['G{}'.format(90 + i - 201)] = lado1
            ws['H{}'.format(90 + i - 201)] = lado2
            ws['I{}'.format(90 + i - 201)] = length
            current_row += 1
        elif i > 240 and i <= 280:
            ws['K{}'.format(90 + i - 241)] = i
            ws['L{}'.format(90 + i - 241)] = lado1
            ws['M{}'.format(90 + i - 241)] = lado2
            ws['N{}'.format(90 + i - 241)] = length
            current_row += 1
        elif i > 280 and i <= 320:
            ws['P{}'.format(90 + i - 281)] = i
            ws['Q{}'.format(90 + i - 281)] = lado1
            ws['R{}'.format(90 + i - 281)] = lado2
            ws['S{}'.format(90 + i - 281)] = length
            current_row += 1

# TABLES LINE 3
        elif i > 320 and i <= 360:
            ws['A{}'.format(133 + i - 321)] = i
            ws['B{}'.format(133 + i - 321)] = lado1
            ws['C{}'.format(133 + i - 321)] = lado2
            ws['D{}'.format(133 + i - 321)] = length
            current_row += 1
        elif i > 360 and i <= 400:
            ws['F{}'.format(133 + i - 361)] = i
            ws['G{}'.format(133 + i - 361)] = lado1
            ws['H{}'.format(133 + i - 361)] = lado2
            ws['I{}'.format(133 + i - 361)] = length
            current_row += 1
        elif i > 400 and i <= 440:
            ws['K{}'.format(133 + i - 401)] = i
            ws['L{}'.format(133 + i - 401)] = lado1
            ws['M{}'.format(133 + i - 401)] = lado2
            ws['N{}'.format(133 + i - 401)] = length
            current_row += 1
        elif i > 440 and i <= 480:
            ws['P{}'.format(133 + i - 441)] = i
            ws['Q{}'.format(133 + i - 441)] = lado1
            ws['R{}'.format(133 + i - 441)] = lado2
            ws['S{}'.format(133 + i - 441)] = length
            current_row += 1

# TABLES LINE 4
        elif i > 480 and i <= 520:
            ws['A{}'.format(176 + i - 481)] = i
            ws['B{}'.format(176 + i - 481)] = lado1
            ws['C{}'.format(176 + i - 481)] = lado2
            ws['D{}'.format(176 + i - 481)] = length
            current_row += 1
        elif i > 520 and i <= 560:
            ws['F{}'.format(176 + i - 521)] = i
            ws['G{}'.format(176 + i - 521)] = lado1
            ws['H{}'.format(176 + i - 521)] = lado2
            ws['I{}'.format(176 + i - 521)] = length
            current_row += 1
        elif i > 560 and i <= 600:
            ws['K{}'.format(176 + i - 561)] = i
            ws['L{}'.format(176 + i - 561)] = lado1
            ws['M{}'.format(176 + i - 561)] = lado2
            ws['N{}'.format(176 + i - 561)] = length
            current_row += 1
        elif i > 600 and i <= 640:
            ws['P{}'.format(176 + i - 601)] = i
            ws['Q{}'.format(176 + i - 601)] = lado1
            ws['R{}'.format(176 + i - 601)] = lado2
            ws['S{}'.format(176 + i - 601)] = length
            current_row += 1

# TABLES LINE 5
        elif i > 640 and i <= 680:
            ws['A{}'.format(219 + i - 641)] = i
            ws['B{}'.format(219 + i - 641)] = lado1
            ws['C{}'.format(219 + i - 641)] = lado2
            ws['D{}'.format(219 + i - 641)] = length
            current_row += 1
        elif i > 680 and i <= 720:
            ws['F{}'.format(219 + i - 681)] = i
            ws['G{}'.format(219 + i - 681)] = lado1
            ws['H{}'.format(219 + i - 681)] = lado2
            ws['I{}'.format(219 + i - 681)] = length
            current_row += 1
        elif i > 720 and i <= 760:
            ws['K{}'.format(219 + i - 721)] = i
            ws['L{}'.format(219 + i - 721)] = lado1
            ws['M{}'.format(219 + i - 721)] = lado2
            ws['N{}'.format(219 + i - 721)] = length
            current_row += 1
        elif i > 760 and i <= 800:
            ws['P{}'.format(219 + i - 761)] = i
            ws['Q{}'.format(219 + i - 761)] = lado1
            ws['R{}'.format(219 + i - 761)] = lado2
            ws['S{}'.format(219 + i - 761)] = length
            current_row += 1

# TABLES LINE 6
        elif i > 800 and i <= 840:
            ws['A{}'.format(262 + i - 801)] = i
            ws['B{}'.format(262 + i - 801)] = lado1
            ws['C{}'.format(262 + i - 801)] = lado2
            ws['D{}'.format(262 + i - 801)] = length
            current_row += 1
        elif i > 840 and i <= 880:
            ws['F{}'.format(262 + i - 841)] = i
            ws['G{}'.format(262 + i - 841)] = lado1
            ws['H{}'.format(262 + i - 841)] = lado2
            ws['I{}'.format(262 + i - 841)] = length
            current_row += 1
        elif i > 880 and i <= 920:
            ws['K{}'.format(262 + i - 881)] = i
            ws['L{}'.format(262 + i - 881)] = lado1
            ws['M{}'.format(262 + i - 881)] = lado2
            ws['N{}'.format(262 + i - 881)] = length
            current_row += 1
        elif i > 920 and i <= 960:
            ws['P{}'.format(262 + i - 921)] = i
            ws['Q{}'.format(262 + i - 921)] = lado1
            ws['R{}'.format(262 + i - 921)] = lado2
            ws['S{}'.format(262 + i - 921)] = length
            current_row += 1


# TABLES LINE 1
        else:
            ws['A{}'.format(current_row)] = data_counter
            ws['B{}'.format(current_row)] = lado1
            ws['C{}'.format(current_row)] = lado2
            ws['D{}'.format(current_row)] = length

            data_counter += 1
            current_row += 1

    # Apply styling
    border = Border(left=Side(style='medium', color='000000'),
                    right=Side(style='medium', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

# TABLES LINE 1
    for col_num in range(1, 5):
        for row_num in range(46, 86 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(6, 10):
        for row_num in range(46, 86 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(11, 15):
        for row_num in range(46, 86 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(16, 20):
        for row_num in range(46, 86 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

# TABLES LINE 2
    for col_num in range(1, 5):
        for row_num in range(89, 129 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(6, 10):
        for row_num in range(89, 129 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(11, 15):
        for row_num in range(89, 129 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(16, 20):
        for row_num in range(89, 129 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border


# TABLE LINE 3
    for col_num in range(1, 5):
        for row_num in range(132, 172 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(6, 10):
        for row_num in range(132, 172 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(11, 15):
        for row_num in range(132, 172 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(16, 20):
        for row_num in range(132, 172 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

# TABLE LINE 4
    for col_num in range(1, 5):  # Columns B to R
        for row_num in range(175, 215 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(6, 10):  # Columns B to R
        for row_num in range(175, 215 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(11, 15):  # Columns B to R
        for row_num in range(175, 215 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(16, 20):  # Columns B to R
        for row_num in range(175, 215 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

# TABLE LINE 5
    for col_num in range(1, 5):  # Columns B to R
        for row_num in range(218, 258 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(6, 10):  # Columns B to R
        for row_num in range(218, 258 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(11, 15):  # Columns B to R
        for row_num in range(218, 258 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(16, 20):  # Columns B to R
        for row_num in range(218, 258 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

# TABLE LINE 6
    for col_num in range(1, 5):  # Columns B to R
        for row_num in range(261, 301 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(6, 10):  # Columns B to R
        for row_num in range(261, 301 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(11, 15):  # Columns B to R
        for row_num in range(261, 301 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(16, 20):  # Columns B to R
        for row_num in range(261, 301 + 1):  # Rows 46 to 87
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)

    # Set the response headers to send the Excel file to the client
    response = make_response(virtual_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=LOTE_BLOQUE_{}.xlsx'.format(
        lote)

    return response

# **************************************************************************************************************************
# *************************************************************TROZA VAL EXCEL*************************************************************
# **************************************************************************************************************************


@app.route('/generate_excelTROZAVal', methods=['POST'])
@login_required
def generate_excelTROZAVal():
    # Retrieve form data from the request
    lote = request.form['lote']
    fecha_ingreso = request.form['fecha_ingreso']
    fecha = request.form['fecha']
    proveedor = request.form['proveedor']
    mp_gorgojo = request.form['mp_gorgojo']
    mp_partida = request.form['mp_partida']
    mp_fisicos = request.form['mp_fisicos']
    mp_quimicos = request.form['mp_quimicos']
    mp_biologico = request.form['mp_biologico']
    cantidadesxlargoTROZA = request.form['cantidadesxlargoTROZA']
    mp_validado_por = request.form['mp_validado_por']

    diferenciam3TROZA = request.form['diferenciam3TROZA']
    diferenciacantidadTROZA = request.form['diferenciacantidadTROZA']
    volm3TROZA = request.form['volm3TROZA']
    CantidadProvTROZA = request.form['CantidadProvTROZA']
    M3Prov = request.form['M3Prov']
    vol_m3_totalTROZA = request.form['vol_m3_totalTROZA']
    unidades_totalTROZA = request.form['unidades_totalTROZA']

    diametroTROZA = request.form['diametroTROZA']
    largoTROZA = request.form['largoTROZA']
    puntaTROZA = request.form['puntaTROZA']
    lado1TROZA = request.form['lado1TROZA']
    lado2TROZA = request.form['lado2TROZA']
    lengthTROZA = request.form['lengthTROZA']

    # Add more fields as needed

    # Create a new workbook
    wb = Workbook()
    ws = wb.active

    # LINEAS DE CUADRICULA DESACTIVADO
    sheet = wb.active
    sheet.sheet_view.showGridLines = False

    ancho_tables = 3
    columna = ws.column_dimensions['D']
    columna.width = ancho_tables
    columna = ws.column_dimensions['N']
    columna.width = ancho_tables

    ancho_id = 4
    columna = ws.column_dimensions['A']
    columna.width = ancho_id
    columna = ws.column_dimensions['E']
    columna.width = ancho_id
    columna = ws.column_dimensions['I']
    columna.width = ancho_id
    columna = ws.column_dimensions['O']
    columna.width = ancho_id

    ancho_largo = 9
    columna = ws.column_dimensions['C']
    columna.width = ancho_largo
    columna = ws.column_dimensions['G']
    columna.width = ancho_largo
    columna = ws.column_dimensions['M']
    columna.width = ancho_largo
    columna = ws.column_dimensions['S']
    columna.width = ancho_largo

    # Write the form data to the worksheet
    ws.merge_cells('C1:Q4')  # Merge cells E1 to J2
    ws['C1'] = 'FORMATO INFORME LIBERADO TROZA'

    # Set alignment and formatting for the merged cell
    merged_cell = ws['C1']
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = Font(bold=True, size=14)

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['C1:Q4']:
        for cell in row:
            cell.border = border

    ws['R1'] = 'Codigo'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R1'].border = border

    ws['S1'] = 'FSAA-VDT-01'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S1'].border = border

    ws['R2'] = 'Vigencia'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R2'].border = border

    ws['S2'] = '21/06/2022'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S2'].border = border

    ws['R3'] = 'Versión'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R3'].border = border

    ws['S3'] = '2'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S3'].border = border

    ws['R4'] = 'Pag'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['R4'].border = border

    ws['S4'] = '1 de 1'
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    ws['S4'].border = border

    font_size = 8
    cell_range = ['S1', 'R1', 'S2', 'R2', 'S3', 'R3', 'S4', 'R4']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)

    ws.merge_cells('C5:D5')
    ws['C5'] = 'LOTE:'
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    for col in range(3, 5):  # Columnas C y D
        for row in range(5, 6):  # Fila 5
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = Font(size=13)

    # Merge y estilo de la celda E5
    ws.merge_cells('E5:F5')
    ws['E5'] = lote
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    for col in range(5, 7):  # Columnas E y F
        for row in range(5, 6):  # Fila 5
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = Font(size=13)

    # Set alignment for the cell containing the lote value
    ws['B1'].alignment = Alignment(horizontal='left')

    # Write other form data to the worksheet
    ws.merge_cells('L7:N7')
    ws['L7'] = 'Fecha Ingreso:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['L7:N7']:
        for cell in row:
            cell.border = border

    ws.merge_cells('O7:P7')
    ws['O7'] = fecha_ingreso
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    ws['O7'].border = border
    for row in ws['O7:P7']:
        for cell in row:
            cell.border = border

    ws.merge_cells('L8:N8')
    ws['L8'] = 'Fecha Validación:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['L8:N8']:
        for cell in row:
            cell.border = border

    ws.merge_cells('O8:P8')
    ws['O8'] = fecha
    border = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(
        style='thin', color='FFFFFF'), bottom=Side(style='thin', color='000000'))
    ws['O8'].border = border
    for row in ws['O8:P8']:
        for cell in row:
            cell.border = border

    ws.merge_cells('D7:F7')
    ws['D7'] = 'Proveedor:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['D7:F7']:
        for cell in row:
            cell.border = border

    ws.merge_cells('G7:J7')
    ws['G7'] = proveedor
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['G7:J7']:
        for cell in row:
            cell.border = border


# *************** CRITERIO DE ACEPTACION Y RECHAZO ******************
    ws.merge_cells('A11:F11')
    ws['A11'] = 'CRITERIO DE ACEPTACION Y RECHAZO'
    ws['A11'].alignment = Alignment(horizontal='center')
    ws['A11'].font = Font(bold=True, size=8)

    # Crear el estilo de borde
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A11:F11']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A12:F12')
    ws['A12'] = 'La madera cumple con las medidas vs remisión?'
    # Crear el estilo de borde
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A12:F12']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A13:F13')
    ws['A13'] = 'La madera cumple con las cantidades vs remisión?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A13:F13']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A14:F14')
    ws['A14'] = 'La madera presenta gorgojo?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A14:F14']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A15:F15')
    ws['A15'] = 'La madera esta partida o quebrada?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A15:F15']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A16:F16')
    ws['A16'] = 'La madera tiene agentes contaminantes físicos?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A16:F16']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A17:F17')
    ws['A17'] = 'La madera tiene agentes contaminantes químicos?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A17:F17']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A18:F18')
    ws['A18'] = 'La madera tiene agentes contaminantes biológicos?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A18:F18']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A19:F19')
    ws['A19'] = 'La madera tiene salvoconducto de movilización?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A19:F19']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A20:F20')
    ws['A20'] = 'Requiere informe no conformidad para el proveedor?'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A20:F20']:
        for cell in row:
            cell.border = border

    font_size = 8
    cell_range = ['A12', 'A13', 'A14', 'A15',
                  'A16', 'A17', 'A18', 'A19', 'A20']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)

    ws.merge_cells('G11:L11')
    ws['G11'] = 'Marque con una (X)'
    ws['G11'].alignment = Alignment(horizontal='center')
    ws['G11'].font = Font(bold=True, size=8)

    # Crear el estilo de borde
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['G11:L11']:
        for cell in row:
            cell.border = border

    ws['G12'] = 'Si'
    ws['G12'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='000000'), bottom=Side(style='thin', color='808080'))
    ws['G12'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H12'].border = border
    ws.merge_cells('H12:I12')

    ws['G13'] = 'Si'
    ws['G13'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G13'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H13'].border = border
    ws.merge_cells('H13:I13')

    ws['G14'] = 'Si'
    ws['G14'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G14'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H14'].border = border
    ws.merge_cells('H14:I14')

    ws['G15'] = 'Si'
    ws['G15'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G15'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H15'].border = border
    ws.merge_cells('H15:I15')

    ws['G16'] = 'Si'
    ws['G16'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G16'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H16'].border = border
    ws.merge_cells('H16:I16')

    ws['G17'] = 'Si'
    ws['G17'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G17'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H17'].border = border
    ws.merge_cells('H17:I17')

    ws['G18'] = 'Si'
    ws['G18'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G18'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H12'].border = border
    ws.merge_cells('H18:I18')

    ws['G19'] = 'Si'
    ws['G19'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['G19'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['H19'].border = border
    ws.merge_cells('H19:I19')

    ws['G20'] = 'Si'
    ws['G20'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='000000'))
    ws['G20'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='000000'))

    ws['H20'].border = border
    ws.merge_cells('H20:I20')

    # _____________NO______________
    ws['J12'] = 'No'
    ws['J12'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J12'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K12'].border = border
    ws.merge_cells('K12:L12')

    ws['J13'] = 'No'
    ws['J13'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J13'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K13'].border = border
    ws.merge_cells('K13:L13')

    ws['J14'] = 'No'
    ws['J14'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J14'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K14'].border = border
    ws.merge_cells('K14:L14')

    ws['J15'] = 'No'
    ws['J15'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J15'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K15'].border = border
    ws.merge_cells('K15:L15')

    ws['J16'] = 'No'
    ws['J16'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J16'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K16'].border = border
    ws.merge_cells('K16:L16')

    ws['J17'] = 'No'
    ws['J17'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J17'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K17'].border = border
    ws.merge_cells('K17:L17')

    ws['J18'] = 'No'
    ws['J18'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J18'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K18'].border = border
    ws.merge_cells('K18:L18')

    ws['J19'] = 'No'
    ws['J19'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
    ws['J19'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'))

    ws['K19'].border = border
    ws.merge_cells('K19:L19')

    ws['J20'] = 'No'
    ws['J20'].alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='808080'), top=Side(
        style='thin', color='808080'), bottom=Side(style='thin', color='000000'))
    ws['J20'].border = border

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='000000'))

    ws['K20'].border = border
    ws.merge_cells('K20:L20')

    font_size = 8
    cell_range = ['G12', 'G13', 'G14', 'G15', 'G16', 'G17', 'G18', 'G19',
                  'G20', 'J12', 'J13', 'J14', 'J15', 'J16', 'J17', 'J18', 'J19', 'J20']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)

    ws.merge_cells('M11:S11')
    ws['M11'] = 'OBSERVACIONES'
    ws['M11'].alignment = Alignment(horizontal='center')
    ws['M11'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in ws['M11:S11']:
        for cell in row:
            cell.border = border

    # _____COMETARIOS______

    ws.merge_cells('M12:S12')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M12:S12']:
        for cell in row:
            cell.border = border

    ws.merge_cells('M13:S13')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M13:S13']:
        for cell in row:
            cell.border = border

    ws['M14'] = mp_gorgojo
    ws.merge_cells('M14:S14')

    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M14:S14']:
        for cell in row:
            cell.border = border

    ws['M15'] = mp_partida
    ws.merge_cells('M15:S15')

    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M15:S15']:
        for cell in row:
            cell.border = border

    ws['M16'] = mp_fisicos
    ws.merge_cells('M16:S16')

    ws.merge_cells('M16:S16')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M16:S16']:
        for cell in row:
            cell.border = border

    ws['M17'] = mp_quimicos
    ws.merge_cells('M17:S17')

    ws.merge_cells('M17:S17')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M17:S17']:
        for cell in row:
            cell.border = border

    ws['M18'] = mp_biologico
    ws.merge_cells('M18:S18')

    ws.merge_cells('M18:S18')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M18:S18']:
        for cell in row:
            cell.border = border

    ws.merge_cells('M19:S19')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='808080'))

    for row in ws['M19:S19']:
        for cell in row:
            cell.border = border

    ws.merge_cells('M20:S20')
    border = Border(left=Side(border_style='thin', color='808080'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='808080'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in ws['M20:S20']:
        for cell in row:
            cell.border = border

    font_size = 8
    cell_range = ['M14', 'M15', 'M16', 'M17', 'M18']
    for cell in cell_range:
        ws[cell].font = Font(size=font_size)
# ______________________________________________________________

    ws.merge_cells('A22:S22')
    ws['A22'] = 'PONDERACION'
    ws['A22'].alignment = Alignment(horizontal='center')
    ws['A22'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in ws['A22:S22']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A23:D24')
    ws['A23'] = 'Agentes Fisícos'
    ws['A23'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A23'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A23:D24']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E23:S24')
    ws['E23'] = 'Los peligros físicos constan de objetos extraños presentes en los alimentos, pudiendo originar lesiones o enfermedades en las personas. Estos peligros se producen por prácticas de inocuidad alimentaria deficientes en cualquier punto de la cadena productiva. Un gran porcentaje de los daños que sufren las personas por el consumo de alimentos y bebidas provienen por los peligros físicos. Gran parte de las lesiones o enfermedades son daños en piezas dentales, molestias gastrointestinales, cortes y quemaduras en la boca y garganta. Los objetos extraños suelen ser: plásticos, piedras, metales, vidrios, cáscaras, maderas, y barros.'
    ws['E23'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['E23'].font = Font(size=6)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E23:S24']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A25:D26')
    ws['A25'] = 'Agentes Quimicos'
    ws['A25'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A25'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A25:D26']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E25:S26')
    ws['E25'] = 'Los peligros químicos se generan de forma natural o se agregan durante cualquier etapa de procesamiento. Los agentes químicos son responsables de enfermedades crónicas, creación de alimentos tóxicos, presencia de sustancias venenosas en los alimentos, entre otros. Estos elementos dañinos ponen en alto riesgo la salud de las personas, por lo que es importante hacer énfasis en el sesion de la materia prima, además de conocer los procesos de producción, cosecha, procesamiento y almacenamiento.'
    ws['E25'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['E25'].font = Font(size=6)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E25:S26']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A27:D28')
    ws['A27'] = 'Agentes Biologicos'
    ws['A27'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A27'].font = Font(bold=True, size=8)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A27:D28']:
        for cell in row:
            cell.border = border

    ws.merge_cells('E27:S28')
    ws['E27'] = 'Entre los 3 tipos de peligros, el peligro biológico es el elemento que representa el mayor riesgo para los humanos. Los riesgos están compuestos por parásitos, bacterias, virus y hongos. Son agentes patógenos que se encuentran generalmente en el ambiente donde estos alimentos se producen. La mayoría de estos agentes son controlados mediante la cocción y una buena manipulación y almacenamiento.'
    ws['E27'].alignment = Alignment(vertical='center', wrap_text=True)
    ws['E27'].font = Font(size=6)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['E27:S28']:
        for cell in row:
            cell.border = border

    ws.merge_cells('C30:E30')
    ws['C30'] = 'Validado por:'
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['C30:E30']:
        for cell in row:
            cell.border = border

    ws.merge_cells('F30:H30')
    ws['F30'] = mp_validado_por
    border = Border(left=Side(border_style='thin', color='FFFFFF'),
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['F30:H30']:
        for cell in row:
            cell.border = border

    ws.merge_cells('A32:B34')
    ws['A32'] = 'Cantidades Por el Largo:'
    ws['A32'].alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    ws['A32'].font = Font(size=10, bold=True)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A32:B34']:
        for cell in row:
            cell.border = border

    ws.merge_cells('C33:S33')
    ws['C33'] = cantidadesxlargoTROZA
    ws['C33'].alignment = Alignment(horizontal='center', vertical='center')

# ______________________________________________________________

    # Set the border style for the range C32:L32
    border_top = Border(top=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=32, max_row=32, min_col=3, max_col=19):
        for cell in row:
            cell.border = border_top

    border_right = Border(right=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=32, max_row=34, min_col=19, max_col=19):
        for cell in row:
            cell.border = border_right

    # Set the border style for the range C34:L34
    border_bottom = Border(bottom=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=34, max_row=34, min_col=3, max_col=19):
        for cell in row:
            cell.border = border_bottom

    cell_l32 = ws['S32']
    border_top_right = Border(top=Side(border_style='thin', color='000000'), right=Side(
        border_style='thin', color='000000'))
    cell_l32.border = border_top_right

    cell_l34 = ws['S34']
    border_bottom_right = Border(bottom=Side(
        border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'))
    cell_l34.border = border_bottom_right

    # Linea sepracion Prov - Lib
    border_left = Border(left=Side(border_style='thin', color='000000'))
    border_right = Border(left=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=8, max_col=8):
        for cell in row:
            cell.border = border_left

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=2, max_col=2):
        for cell in row:
            cell.border = border_right

    DiffT = PatternFill(start_color="D3D3D3",
                        end_color="D3D3D3", fill_type="solid")
    ws.merge_cells('I38:J38')
    ws['I38'] = 'Dif. Vol-M3:'
    ws['I38'].alignment = Alignment(horizontal='center')
    ws['I38'].font = Font(bold=True, size=10)
    ws['I38'].fill = DiffT

    ws['K38'] = diferenciam3TROZA
    ws['K38'].fill = DiffT

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=3, max_col=7):
        for cell in row:
            cell.border = border_bottom

    thick_border = Border(left=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    # Aplica el borde grueso a las celdas I38 a L38
    for row in ws.iter_rows(min_row=38, max_row=38, min_col=9, max_col=10):
        for cell in row:
            cell.border = thick_border

    thick_border = Border(right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))
    for row in ws.iter_rows(min_row=38, max_row=38, min_col=11, max_col=11):
        for cell in row:
            cell.border = thick_border

    DiffT = PatternFill(start_color="D3D3D3",
                        end_color="D3D3D3", fill_type="solid")
    ws.merge_cells('I40:J40')
    ws['I40'] = 'Dif. Unidades:'
    ws['I40'].alignment = Alignment(horizontal='center')
    ws['I40'].font = Font(bold=True, size=10)
    ws['I40'].fill = DiffT  # Aplica el relleno gris a la celda K33

    ws['K40'] = diferenciacantidadTROZA
    ws['K40'].fill = DiffT

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=3, max_col=7):
        for cell in row:
            cell.border = border_bottom

    thick_border = Border(left=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    # Aplica el borde grueso a las celdas I38 a L38
    for row in ws.iter_rows(min_row=40, max_row=40, min_col=9, max_col=10):
        for cell in row:
            cell.border = thick_border

    thick_border = Border(right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))
    for row in ws.iter_rows(min_row=40, max_row=40, min_col=11, max_col=11):
        for cell in row:
            cell.border = thick_border

    ws.merge_cells('B36:G36')
    ws['B36'] = proveedor
    ws['B36'].alignment = Alignment(horizontal='center')
    ws['B36'].font = Font(bold=True, size=11)
    border_bottom = Border(bottom=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=2, max_col=7):
        for cell in row:
            cell.border = border_bottom

    ws.merge_cells('C38:E38')
    ws['C38'] = 'M3 Remisión:'
    ws['C38'].alignment = Alignment(horizontal='center')
    ws['C38'].font = Font(size=10)
    ws['F38'] = M3Prov

    ws.merge_cells('C40:E40')
    ws['C40'] = 'Unds Remisión:'
    ws['C40'].alignment = Alignment(horizontal='center')
    ws['C40'].font = Font(size=10)
    ws['F40'] = CantidadProvTROZA

    ws.merge_cells('C42:E42')
    ws['C42'] = 'Vol-M3 Tabla:'
    ws['C42'].alignment = Alignment(horizontal='center')
    ws['C42'].font = Font(size=10)
    ws['F42'] = volm3TROZA

    ws.merge_cells('M36:R36')
    ws['M36'] = 'MADERAS INDSUTRIALES S.A.S'
    ws['M36'].alignment = Alignment(horizontal='center')
    ws['M36'].font = Font(bold=True, size=11)
    border_bottom = Border(bottom=Side(border_style='thin', color='000000'))

    for row in ws.iter_rows(min_row=36, max_row=36, min_col=13, max_col=18):
        for cell in row:
            cell.border = border_bottom

    ws.merge_cells('N38:P38')
    ws['N38'] = 'Vol-M3 Total:'
    ws['N38'].alignment = Alignment(horizontal='center')
    ws['N38'].font = Font(size=10)

    ws.merge_cells('Q38:R38')
    ws['Q38'] = vol_m3_totalTROZA
    ws['Q38'].font = Font(size=10)

    ws.merge_cells('N40:P40')
    ws['N40'] = 'Unds Total:'
    ws['N40'].alignment = Alignment(horizontal='center')
    ws['N40'].font = Font(size=10)

    ws.merge_cells('Q40:R40')
    ws['Q40'] = unidades_totalTROZA
    ws['Q40'].font = Font(size=10)

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=19, max_col=19):
        for cell in row:
            cell.border = border_left

    for row in ws.iter_rows(min_row=37, max_row=42, min_col=13, max_col=13):
        for cell in row:
            cell.border = border_right


# ______________________________________________________________
    # Add the image to cell D1
    img = Image('src/static/img/splash.png')
    img.width = 80
    img.height = 80
    img.left = 40  # Agregar un margen izquierdo de 40 unidades
    ws.add_image(img, 'A1')
    ws.merge_cells('A1:B4')

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Aplicar el estilo de borde a la celda
    for row in ws['A1:B4']:
        for cell in row:
            cell.border = border

    # ******* TABLE PROVEEDOR ******

    # TABLE LINE 1
    ws.merge_cells('A45:G45')
    ws['A45'] = proveedor
    ws['A45'].alignment = Alignment(horizontal='center')
    ws['A45'].font = Font(bold=True, size=11)

    # TABLE LINE 2
    ws.merge_cells('A88:G88')
    ws['A88'] = proveedor
    ws['A88'].alignment = Alignment(horizontal='center')
    ws['A88'].font = Font(bold=True, size=11)

    # TABLE LINE 3
    ws.merge_cells('A131:G131')
    ws['A131'] = proveedor
    ws['A131'].alignment = Alignment(horizontal='center')
    ws['A131'].font = Font(bold=True, size=11)

    # TABLE LINE 4
    ws.merge_cells('A174:G174')
    ws['A174'] = proveedor
    ws['A174'].alignment = Alignment(horizontal='center')
    ws['A174'].font = Font(bold=True, size=11)

    # TABLE LINE 5
    ws.merge_cells('A217:G217')
    ws['A217'] = proveedor
    ws['A217'].alignment = Alignment(horizontal='center')
    ws['A217'].font = Font(bold=True, size=11)

    header_font = Font(size=8, bold=True)
    ws['A46'] = 'ID'
    ws['B46'] = 'Dia/Menor'
    ws['C46'] = 'Largo'

    # TABLE LINE 1
    ws['A46'].font = header_font
    ws['B46'].font = header_font
    ws['C46'].font = header_font

    ws['E46'].font = header_font
    ws['F46'].font = header_font
    ws['G46'].font = header_font

    # TABLE LINE 2
    ws['A89'].font = header_font
    ws['B89'].font = header_font
    ws['C89'].font = header_font

    ws['E89'].font = header_font
    ws['F89'].font = header_font
    ws['G89'].font = header_font

    # TABLE LINE 3
    ws['A132'].font = header_font
    ws['B132'].font = header_font
    ws['C132'].font = header_font

    ws['E132'].font = header_font
    ws['F132'].font = header_font
    ws['G132'].font = header_font

    # TABLE LINE 4
    ws['A175'].font = header_font
    ws['B175'].font = header_font
    ws['C175'].font = header_font

    ws['E175'].font = header_font
    ws['F175'].font = header_font
    ws['G175'].font = header_font

    # TABLE LINE 5
    ws['A218'].font = header_font
    ws['B218'].font = header_font
    ws['C218'].font = header_font

    ws['E218'].font = header_font
    ws['F218'].font = header_font
    ws['G218'].font = header_font

    # Split the diametroTROZA and largoTROZA strings into lists
    diametroTROZAArray = diametroTROZA.split(',')
    largoTROZAArray = largoTROZA.split(',')

    max_rows_per_cell = 40  # Maximum rows per cell
    current_row = 47  # Starting row for data
    data_counter = 1

    for i, (diametro, largoT) in enumerate(zip(diametroTROZAArray, largoTROZAArray), start=1):
        diametro = float(diametro.strip())
        largoT = float(largoT.strip())
        if i % max_rows_per_cell == 1 and i > max_rows_per_cell:
            current_row = 47  # Reset to the starting row of the next cell
            ws['E46'] = 'ID'
            ws['F46'] = 'Dia/Menor'
            ws['G46'] = 'Largo'

            ws['A89'] = 'ID'
            ws['B89'] = 'Dia/Menor'
            ws['C89'] = 'Largo'
            ws['E89'] = 'ID'
            ws['F89'] = 'Dia/Menor'
            ws['G89'] = 'Largo'

            ws['A132'] = 'ID'
            ws['B132'] = 'Dia/Menor'
            ws['C132'] = 'Largo'
            ws['E132'] = 'ID'
            ws['F132'] = 'Dia/Menor'
            ws['G132'] = 'Largo'

            ws['A175'] = 'ID'
            ws['B175'] = 'Dia/Menor'
            ws['C175'] = 'Largo'
            ws['E175'] = 'ID'
            ws['F175'] = 'Dia/Menor'
            ws['G175'] = 'Largo'

            ws['A218'] = 'ID'
            ws['B218'] = 'Dia/Menor'
            ws['C218'] = 'Largo'
            ws['E218'] = 'ID'
            ws['F218'] = 'Dia/Menor'
            ws['G218'] = 'Largo'

        # TABLES LINE 1
        if i > max_rows_per_cell and i <= 80:
            ws['E{}'.format(current_row)] = i
            ws['F{}'.format(current_row)] = diametro
            ws['G{}'.format(current_row)] = largoT

            current_row += 1

        # TABLES LINE 2
        elif i > 80 and i <= 120:
            ws['A{}'.format(90 + i - 81)] = i
            ws['B{}'.format(90 + i - 81)] = diametro
            ws['C{}'.format(90 + i - 81)] = largoT

            current_row += 1
        elif i > 120 and i <= 160:
            ws['E{}'.format(90 + i - 121)] = i
            ws['F{}'.format(90 + i - 121)] = diametro
            ws['G{}'.format(90 + i - 121)] = largoT

            current_row += 1

        # TABLES LINE 3
        elif i > 160 and i <= 200:
            ws['A{}'.format(133 + i - 161)] = i
            ws['B{}'.format(133 + i - 161)] = diametro
            ws['C{}'.format(133 + i - 161)] = largoT

            current_row += 1
        elif i > 200 and i <= 240:
            ws['E{}'.format(133 + i - 201)] = i
            ws['F{}'.format(133 + i - 201)] = diametro
            ws['G{}'.format(133 + i - 201)] = largoT

            current_row += 1

        # TABLES LINE 4
        elif i > 240 and i <= 280:
            ws['A{}'.format(176 + i - 241)] = i
            ws['B{}'.format(176 + i - 241)] = diametro
            ws['C{}'.format(176 + i - 241)] = largoT

            current_row += 1
        elif i > 280 and i <= 320:
            ws['E{}'.format(176 + i - 281)] = i
            ws['F{}'.format(176 + i - 281)] = diametro
            ws['G{}'.format(176 + i - 281)] = largoT

            current_row += 1

        # TABLES LINE 5
        elif i > 320 and i <= 360:
            ws['A{}'.format(219 + i - 321)] = i
            ws['B{}'.format(219 + i - 321)] = diametro
            ws['C{}'.format(219 + i - 321)] = largoT

            current_row += 1
        elif i > 360 and i <= 400:
            ws['E{}'.format(219 + i - 361)] = i
            ws['F{}'.format(219 + i - 361)] = diametro
            ws['G{}'.format(219 + i - 361)] = largoT

            current_row += 1

        # TABLES LINE 1
        else:
            ws['A{}'.format(current_row)] = data_counter
            ws['B{}'.format(current_row)] = diametro
            ws['C{}'.format(current_row)] = largoT

            data_counter += 1
            current_row += 1

    # Apply styling
    border = Border(left=Side(style='medium', color='000000'),  # Medium weight and black color
                    # Medium weight and black color
                    right=Side(style='medium', color='000000'),
                    # Thin weight and black color
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))  # Thin weight and black color

    # TABLES LINE 1
    for col_num in range(1, 4):
        for row_num in range(46, 86 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(5, 8):
        for row_num in range(46, 86 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # TABLES LINE 2
    for col_num in range(1, 4):
        for row_num in range(89, 129 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(5, 8):
        for row_num in range(89, 129 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # TABLES LINE 3
    for col_num in range(1, 4):
        for row_num in range(132, 172 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(5, 8):
        for row_num in range(132, 172 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # TABLES LINE 4
    for col_num in range(1, 4):
        for row_num in range(175, 215 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(5, 8):
        for row_num in range(175, 215 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # TABLES LINE 5
    for col_num in range(1, 4):
        for row_num in range(218, 258 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(5, 8):
        for row_num in range(218, 258 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # ******* TABLE LIBERADO ******

    # TABLE LINE 1
    ws.merge_cells('I45:S45')
    ws['I45'] = "MADERAS INDSUTRIALES S.A.S"
    ws['I45'].alignment = Alignment(horizontal='center')
    ws['I45'].font = Font(bold=True, size=11)

    # TABLE LINE 2
    ws.merge_cells('I88:S88')
    ws['I88'] = "MADERAS INDSUTRIALES S.A.S"
    ws['I88'].alignment = Alignment(horizontal='center')
    ws['I88'].font = Font(bold=True, size=11)

    # TABLE LINE 3
    ws.merge_cells('I131:S131')
    ws['I131'] = "MADERAS INDSUTRIALES S.A.S"
    ws['I131'].alignment = Alignment(horizontal='center')
    ws['I131'].font = Font(bold=True, size=11)

    # TABLE LINE 4
    ws.merge_cells('I174:S174')
    ws['I174'] = "MADERAS INDSUTRIALES S.A.S"
    ws['I174'].alignment = Alignment(horizontal='center')
    ws['I174'].font = Font(bold=True, size=11)

    # TABLE LINE 5
    ws.merge_cells('I217:S217')
    ws['I217'] = "MADERAS INDSUTRIALES S.A.S"
    ws['I217'].alignment = Alignment(horizontal='center')
    ws['I217'].font = Font(bold=True, size=11)

    header_font = Font(size=8, bold=True)
    ws['I46'] = 'ID'
    ws['J46'] = 'Punta'
    ws['K46'] = 'Lado 1'
    ws['L46'] = 'Lado 2'
    ws['M46'] = 'Largo'

    # TABLE LINE 1
    ws['I46'].font = header_font
    ws['J46'].font = header_font
    ws['K46'].font = header_font
    ws['L46'].font = header_font
    ws['M46'].font = header_font

    ws['O46'].font = header_font
    ws['P46'].font = header_font
    ws['Q46'].font = header_font
    ws['R46'].font = header_font
    ws['S46'].font = header_font

    # TABLE LINE 2
    ws['I89'].font = header_font
    ws['J89'].font = header_font
    ws['K89'].font = header_font
    ws['L89'].font = header_font
    ws['M89'].font = header_font

    ws['O89'].font = header_font
    ws['P89'].font = header_font
    ws['Q89'].font = header_font
    ws['R89'].font = header_font
    ws['S89'].font = header_font

    # TABLE LINE 3
    ws['I132'].font = header_font
    ws['J132'].font = header_font
    ws['K132'].font = header_font
    ws['L132'].font = header_font
    ws['M132'].font = header_font

    ws['O132'].font = header_font
    ws['P132'].font = header_font
    ws['Q132'].font = header_font
    ws['R132'].font = header_font
    ws['S132'].font = header_font

    # TABLE LINE 4
    ws['I175'].font = header_font
    ws['J175'].font = header_font
    ws['K175'].font = header_font
    ws['L175'].font = header_font
    ws['M175'].font = header_font

    ws['O175'].font = header_font
    ws['P175'].font = header_font
    ws['Q175'].font = header_font
    ws['R175'].font = header_font
    ws['S175'].font = header_font

    # TABLE LINE 5
    ws['I218'].font = header_font
    ws['J218'].font = header_font
    ws['K218'].font = header_font
    ws['L218'].font = header_font
    ws['M218'].font = header_font

    ws['O218'].font = header_font
    ws['P218'].font = header_font
    ws['Q218'].font = header_font
    ws['R218'].font = header_font
    ws['S218'].font = header_font

    puntaTROZAArray = puntaTROZA.split(',')
    lado1TROZAArray = lado1TROZA.split(',')
    lado2TROZAArray = lado2TROZA.split(',')
    lengthTROZAArray = lengthTROZA.split(',')

    max_rows_per_cell = 40  # Maximum rows per cell
    current_row = 47  # Starting row for data
    data_counter = 1

    for i, (punta, lado1B, lado2B, LargoB) in enumerate(zip(puntaTROZAArray, lado1TROZAArray, lado2TROZAArray, lengthTROZAArray), start=1):
        # Remove any leading/trailing spaces and convert to float
        punta = float(punta.strip())
        lado1B = float(lado1B.strip())
        # Remove any leading/trailing spaces and convert to float
        lado2B = float(lado2B.strip())
        LargoB = float(LargoB.strip())
        if i % max_rows_per_cell == 1 and i > max_rows_per_cell:
            current_row = 47  # Reset to the starting row of the next cell
            ws['O46'] = 'ID'
            ws['P46'] = 'Punta'
            ws['Q46'] = 'Lado 1'
            ws['R46'] = 'Lado 2'
            ws['S46'] = 'Largo'

            ws['I89'] = 'ID'
            ws['J89'] = 'Punta'
            ws['K89'] = 'Lado 1'
            ws['L89'] = 'Lado 2'
            ws['M89'] = 'Largo'
            ws['O89'] = 'ID'
            ws['P89'] = 'Punta'
            ws['Q89'] = 'Lado 1'
            ws['R89'] = 'Lado 2'
            ws['S89'] = 'Largo'

            ws['I132'] = 'ID'
            ws['J132'] = 'Punta'
            ws['K132'] = 'Lado 1'
            ws['L132'] = 'Lado 2'
            ws['M132'] = 'Largo'
            ws['O132'] = 'ID'
            ws['P132'] = 'Punta'
            ws['Q132'] = 'Lado 1'
            ws['R132'] = 'Lado 2'
            ws['S132'] = 'Largo'

            ws['I175'] = 'ID'
            ws['J175'] = 'Punta'
            ws['K175'] = 'Lado 1'
            ws['L175'] = 'Lado 2'
            ws['M175'] = 'Largo'
            ws['O175'] = 'ID'
            ws['P175'] = 'Punta'
            ws['Q175'] = 'Lado 1'
            ws['R175'] = 'Lado 2'
            ws['S175'] = 'Largo'

            ws['I218'] = 'ID'
            ws['J218'] = 'Punta'
            ws['K218'] = 'Lado 1'
            ws['L218'] = 'Lado 2'
            ws['M218'] = 'Largo'
            ws['O218'] = 'ID'
            ws['P218'] = 'Punta'
            ws['Q218'] = 'Lado 1'
            ws['R218'] = 'Lado 2'
            ws['S218'] = 'Largo'

        # TABLES LINE 1
        if i > max_rows_per_cell and i <= 80:
            ws['O{}'.format(current_row)] = i
            ws['P{}'.format(current_row)] = punta
            ws['Q{}'.format(current_row)] = lado1B
            ws['R{}'.format(current_row)] = lado2B
            ws['S{}'.format(current_row)] = LargoB

            current_row += 1

        # TABLES LINE 2
        elif i > 80 and i <= 120:
            ws['I{}'.format(90 + i - 81)] = i
            ws['J{}'.format(90 + i - 81)] = punta
            ws['K{}'.format(90 + i - 81)] = lado1B
            ws['L{}'.format(90 + i - 81)] = lado2B
            ws['M{}'.format(90 + i - 81)] = LargoB

            current_row += 1
        elif i > 120 and i <= 160:
            ws['O{}'.format(90 + i - 121)] = i
            ws['P{}'.format(90 + i - 121)] = punta
            ws['Q{}'.format(90 + i - 121)] = lado1B
            ws['R{}'.format(90 + i - 121)] = lado2B
            ws['S{}'.format(90 + i - 121)] = LargoB

            current_row += 1

         # TABLES LINE 3
        elif i > 160 and i <= 200:
            ws['I{}'.format(133 + i - 161)] = i
            ws['J{}'.format(133 + i - 161)] = punta
            ws['K{}'.format(133 + i - 161)] = lado1B
            ws['L{}'.format(133 + i - 161)] = lado2B
            ws['M{}'.format(133 + i - 161)] = LargoB

            current_row += 1
        elif i > 200 and i <= 240:
            ws['O{}'.format(133 + i - 201)] = i
            ws['P{}'.format(133 + i - 201)] = punta
            ws['Q{}'.format(133 + i - 201)] = lado1B
            ws['R{}'.format(133 + i - 201)] = lado2B
            ws['S{}'.format(133 + i - 201)] = LargoB

            current_row += 1

        # TABLES LINE 4
        elif i > 240 and i <= 280:
            ws['I{}'.format(176 + i - 241)] = i
            ws['J{}'.format(176 + i - 241)] = punta
            ws['K{}'.format(176 + i - 241)] = lado1B
            ws['L{}'.format(176 + i - 241)] = lado2B
            ws['M{}'.format(176 + i - 241)] = LargoB

            current_row += 1
        elif i > 280 and i <= 320:
            ws['O{}'.format(176 + i - 281)] = i
            ws['P{}'.format(176 + i - 281)] = punta
            ws['Q{}'.format(176 + i - 281)] = lado1B
            ws['R{}'.format(176 + i - 281)] = lado2B
            ws['S{}'.format(176 + i - 281)] = LargoB

            current_row += 1

        # TABLES LINE 5
        elif i > 320 and i <= 360:
            ws['I{}'.format(219 + i - 321)] = i
            ws['J{}'.format(219 + i - 321)] = punta
            ws['K{}'.format(219 + i - 321)] = lado1B
            ws['L{}'.format(219 + i - 321)] = lado2B
            ws['M{}'.format(219 + i - 321)] = LargoB

            current_row += 1
        elif i > 360 and i <= 400:
            ws['O{}'.format(219 + i - 361)] = i
            ws['P{}'.format(219 + i - 361)] = punta
            ws['Q{}'.format(219 + i - 361)] = lado1B
            ws['R{}'.format(219 + i - 361)] = lado2B
            ws['S{}'.format(219 + i - 361)] = LargoB

            current_row += 1

        # TABLES LINE 1
        else:
            ws['I{}'.format(current_row)] = data_counter
            ws['J{}'.format(current_row)] = punta
            ws['K{}'.format(current_row)] = lado1B
            ws['L{}'.format(current_row)] = lado2B
            ws['M{}'.format(current_row)] = LargoB

            data_counter += 1
            current_row += 1

    # Apply styling
    border = Border(left=Side(style='medium', color='000000'),  # Medium weight and black color
                    # Medium weight and black color
                    right=Side(style='medium', color='000000'),
                    # Thin weight and black color
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))  # Thin weight and black color

    # TABLES LINE 1
    for col_num in range(9, 14):
        for row_num in range(46, 86 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(15, 20):
        for row_num in range(46, 86 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # TABLES LINE 2
    for col_num in range(9, 14):
        for row_num in range(89, 129 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(15, 20):
        for row_num in range(89, 129 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # TABLES LINE 3
    for col_num in range(9, 14):
        for row_num in range(132, 172 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(15, 20):
        for row_num in range(132, 172 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # TABLES LINE 4
    for col_num in range(9, 14):
        for row_num in range(175, 215 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(15, 20):
        for row_num in range(175, 215 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    # TABLES LINE 5
    for col_num in range(9, 14):
        for row_num in range(218, 258 + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
    for col_num in range(15, 20):
        for row_num in range(218, 258 + 1):  # Rows 46 to 129
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border

    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)

    # Set the response headers to send the Excel file to the client
    response = make_response(virtual_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=LOTE_TROZA_{}.xlsx'.format(
        lote)

    return response


@app.route('/login', methods=['GET', 'POST'])
def login():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        # Perform user authentication and retrieve user data from the database
        response = requests.get(
            'http://oficial2024.maderasindustrialesoftware.com/admin_user/loginWEB.php')
        if response.status_code == 200:
            users = response.json()

            logged_user = None
            for user in users:
                if user['username'] == username:
                    logged_user = User(
                        user['id'], user['username'], user['password'], user['fullname'], user['rol'])
                    break

            if logged_user:
                stored_password = logged_user.password
                if bcrypt.checkpw(password.encode('utf-8'), stored_password.encode('utf-8')):
                    response = requests.post(
                        'http://oficial2024.maderasindustrialesoftware.com/admin_user/loginWEB.php', json=logged_user.to_dict())
                    if response.status_code == 200:
                        # User session updated successfully in the database
                        login_user(logged_user)
                        return redirect(url_for('home'))
                    else:
                        flash(
                            "Error al actualizar la sesión del usuario en la base de datos.")
                else:
                    flash("Contraseña inválida...")
            else:
                flash("Usuario no encontrado...")
        else:
            flash("Error al obtener los datos del usuario desde la base de datos.")

    return render_template('auth/login.html')


@app.route('/logout')
def logout():
    if current_user.is_authenticated:
        logout_user()
    return redirect(url_for('login'))


# OBTENCION DE USUARIO DE LA BD
@app.route('/admin_user')
@login_required
def admin_user():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')

    if current_user.rol != 'ADMIN':
        return redirect(url_for('home'))  # Redirect to a suitable page

    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/admin_user/get_usersADMIN.php')
    user = response.json()
    return render_template('admin_html/admin_user.html', users=user)


# CREACION DE USUARIOS
@app.route('/agregar_user', methods=['GET', 'POST'])
@login_required
def agregar_user():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/admin_user/add_user.php')
    username_exist = response.json()

    if request.method == 'POST':
        # Obtener datos del formulario
        username = request.form['username']
        password = request.form['password']
        fullname = request.form['fullname']
        rol = request.form['rol']

        if username in username_exist['usernames']:
            flash("El Usuario ingresado ya existe.")
            return redirect(url_for('agregar_user'))

        # Hash the password
        hashed_password = bcrypt.hashpw(
            password.encode('utf-8'), bcrypt.gensalt())

        data = {
            'id': '',
            'username': username,
            'password': hashed_password.decode('utf-8'),
            'fullname': fullname,
            'rol': rol,
        }

        response = requests.post(
            'http://oficial2024.maderasindustrialesoftware.com/admin_user/add_user.php', data=data)

        if response.status_code == 200:
            return redirect(url_for('admin_user'))
        else:
            flash("Error al crear el Usuario.")
    else:
        # Renderizar el formulario con el número de lote siguiente
        return render_template('admin_html/agregar_user.html')


@app.route('/cambiar_user', methods=['POST'])
@login_required
def cambiar_user():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    try:
        if request.method == 'POST':
            # Obtener datos del formulario
            id = request.form['id']
            username = request.form['username']
            password = request.form['password']
            fullname = request.form['fullname']
            rol = request.form['rol']

            # Hash the password
            hashed_password = bcrypt.hashpw(
                password.encode('utf-8'), bcrypt.gensalt())

            data = {
                'id': id,
                'username': username,
                'password': hashed_password.decode('utf-8'),
                'fullname': fullname,
                'rol': rol,
            }

            # Imprimir los datos antes de la solicitud HTTP
            print("Datos enviados:")
            print("ID:", id)
            print("Username:", username)
            print("Password:", password)
            print("Fullname:", fullname)
            print("Rol:", rol)

            response = requests.post(
                'http://oficial2024.maderasindustrialesoftware.com/admin_user/editar_user.php', data=data)

            if response.status_code == 200:
                # Redirect to appropriate page
                return redirect(url_for('admin_user'))
            else:
                flash("Error al actualizar el Usuario.")

    except Exception as ex:
        flash('Ocurrió un error al actualizar el usuario: {}'.format(ex))

    # Redirect to appropriate page
    return redirect(url_for('admin_html/admin_user'))



@app.route('/editar_user')
@login_required
def informacion_usuario():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    user_id = request.args.get('id')

    # Obtener los datos del usuario de la base de datos
    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/admin_user/delete.php')
    users = response.json()

    # Buscar el usuario correspondiente al ID solicitado
    user = None
    for u in users:
        if u['id'] == user_id:
            user = u
            break

    # Renderizar la plantilla 'editar_user.html' y pasar los datos del usuario como argumentos
    return render_template('admin_html/editar_user.html', user=user)


@app.route('/DeshabilitarUser', methods=['POST'])
@login_required
def DeshabilitarUser():
    try:
        user_id = request.json['id']
        response = requests.post(
            'http://oficial2024.maderasindustrialesoftware.com/admin_user/delete.php', data={'id': user_id})
        return jsonify({'message': response.json()['message']}), response.status_code
    except Exception as ex:
        return jsonify({'error': str(ex)}), 500


@app.route('/home')
@login_required
def home():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    else:
        return render_template('home.html')


@app.route('/info_entrada')
@login_required
def info_entrada():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    # INFO PP GET
    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/get_infoPP.php')
    lotePP = response.json()

# INFO BLOQUE GET
    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/get_infoBLOQUE.php')
    loteBLOQUE = response.json()


# INFO TROZA GET
    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/get_infoTROZA.php')
    loteTROZA = response.json()

    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/last_lote.php')
    lote = response.text.strip()

    return render_template('info_entrada_html/info_entrada.html', usersPP=lotePP, usersBLOQUE=loteBLOQUE, usersTROZA=loteTROZA, last_infoLOTE=lote)


@app.route('/agregar_mp')
@login_required
def agregar_mp():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    return render_template('info_entrada_html/agregar_mp.html')


@app.route('/agregar_infoPP', methods=['GET', 'POST'])
@login_required
def agregar_infoPP():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    # Obtener last_lote desde la URL
    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/lotes_generate.php')
    generate_lote = response.text.strip()

    if request.method == 'POST':
        # Obtener datos del formulario
        lotePP = request.form['lotePP']
        fecha_ingresoPP = request.form['fecha_ingresoPP']
        proveedorPP = request.form['proveedorPP']
        receptorPP = request.form['receptorPP']
        remision_provPP = request.form['remision_provPP']
        transportadorPP = request.form['transportadorPP']
        placaPP = request.form['placaPP']
        presentacionPP = request.form['presentacionPP']
        m3PP = request.form['m3PP']
        pulgPP = request.form['pulgPP']
        cantidadtotalPP = request.form['cantidadtotalPP']
        infotablepp = request.form['infotablepp']

        data = {
            'lotePP': lotePP,
            'fecha_ingresoPP': fecha_ingresoPP,
            'proveedorPP': proveedorPP,
            'receptorPP': receptorPP,
            'remision_provPP': remision_provPP,
            'transportadorPP': transportadorPP,
            'placaPP': placaPP,
            'presentacionPP': presentacionPP,
            'm3PP': m3PP,
            'pulgPP': pulgPP,
            'cantidadtotalPP': cantidadtotalPP,
            'infotablepp': infotablepp,
        }

        response = requests.post(
            'http://oficial2024.maderasindustrialesoftware.com/info_entrada/informePP.php', data=data)

        if response.status_code == 200:
            return redirect(url_for('info_entrada'))
        else:
            flash("Error al crear el informe.")
    else:
        # Renderizar el formulario con el número de lote siguiente
        return render_template('info_entrada_html/agregar_infoPP.html', lote=generate_lote)


@app.route('/agregar_infoBLOQUE', methods=['GET', 'POST'])
@login_required
def agregar_infoBLOQUE():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    # Obtener last_lote desde la URL
    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/lotes_generate.php')
    generate_lote = response.text.strip()

    if request.method == 'POST':
        # Obtener datos del formulario
        loteBLOQUE = request.form['loteBLOQUE']
        fecha_ingresoBLOQUE = request.form['fecha_ingresoBLOQUE']
        proveedorBLOQUE = request.form['proveedorBLOQUE']
        receptorBLOQUE = request.form['receptorBLOQUE']
        remision_provBLOQUE = request.form['remision_provBLOQUE']
        talo_salvoconductoBLOQUE = request.form['talo_salvoconductoBLOQUE']
        remi_salvoconductoBLOQUE = request.form['remi_salvoconductoBLOQUE']
        nombre_razonBLOQUE = request.form['nombre_razonBLOQUE']
        entidadBLOQUE = request.form['entidadBLOQUE']
        no_registroBLOQUE = request.form['no_registroBLOQUE']
        departamentoBLOQUE = request.form['departamentoBLOQUE']
        municipioBLOQUE = request.form['municipioBLOQUE']
        veredaBLOQUE = request.form['veredaBLOQUE']
        transportadorBLOQUE = request.form['transportadorBLOQUE']
        placaBLOQUE = request.form['placaBLOQUE']
        comunBLOQUE = request.form['comunBLOQUE']
        cientificoBLOQUE = request.form['cientificoBLOQUE']
        presentacionBLOQUE = request.form['presentacionBLOQUE']
        pulgadasBLOQUE = request.form['pulgadasBLOQUE']
        infotablebloque = request.form['infotablebloque']


        data = {
            'loteBLOQUE': loteBLOQUE,
            'fecha_ingresoBLOQUE': fecha_ingresoBLOQUE,
            'proveedorBLOQUE': proveedorBLOQUE,
            'receptorBLOQUE': receptorBLOQUE,
            'remision_provBLOQUE': remision_provBLOQUE,
            'talo_salvoconductoBLOQUE': talo_salvoconductoBLOQUE,
            'remi_salvoconductoBLOQUE': remi_salvoconductoBLOQUE,
            'nombre_razonBLOQUE': nombre_razonBLOQUE,
            'entidadBLOQUE': entidadBLOQUE,
            'no_registroBLOQUE': no_registroBLOQUE,
            'departamentoBLOQUE': departamentoBLOQUE,
            'municipioBLOQUE': municipioBLOQUE,
            'veredaBLOQUE': veredaBLOQUE,
            'transportadorBLOQUE': transportadorBLOQUE,
            'placaBLOQUE': placaBLOQUE,
            'comunBLOQUE': comunBLOQUE,
            'cientificoBLOQUE': cientificoBLOQUE,
            'presentacionBLOQUE': presentacionBLOQUE,
            'pulgadasBLOQUE': pulgadasBLOQUE,
            'infotablebloque': infotablebloque,
        }

        response = requests.post(
            'http://oficial2024.maderasindustrialesoftware.com/info_entrada/informeBLOQUE.php', data=data)

        if response.status_code == 200:
            return redirect(url_for('info_entrada'))
        else:
            flash("Error al crear el informe.")
    else:
        # Renderizar el formulario con el número de lote siguiente
        return render_template('info_entrada_html/agregar_infoBLOQUE.html', lote=generate_lote)


@app.route('/agregar_infoTROZA', methods=['GET', 'POST'])
@login_required
def agregar_infoTROZA():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    # Obtener last_lote desde la URL
    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/lotes_generate.php')
    generate_lote = response.text.strip()

    if request.method == 'POST':
        # Obtener datos del formulario
        loteTROZA = request.form['loteTROZA']
        fecha_ingresoTROZA = request.form['fecha_ingresoTROZA']
        proveedorTROZA = request.form['proveedorTROZA']
        receptorTROZA = request.form['receptorTROZA']
        remision_provTROZA = request.form['remision_provTROZA']
        talo_salvoconductoTROZA = request.form['talo_salvoconductoTROZA']
        remi_salvoconductoTROZA = request.form['remi_salvoconductoTROZA']
        nombre_razonTROZA = request.form['nombre_razonTROZA']
        entidadTROZA = request.form['entidadTROZA']
        no_registroTROZA = request.form['no_registroTROZA']
        departamentoTROZA = request.form['departamentoTROZA']
        municipioTROZA = request.form['municipioTROZA']
        veredaTROZA = request.form['veredaTROZA']
        transportadorTROZA = request.form['transportadorTROZA']
        placaTROZA = request.form['placaTROZA']
        comunTROZA = request.form['comunTROZA']
        cientificoTROZA = request.form['cientificoTROZA']
        presentacionTROZA = request.form['presentacionTROZA']
        infotabletroza = request.form['infotabletroza']
        volm3TROZA = request.form['volm3TROZA']
        M3ProvTROZA = request.form['M3ProvTROZA']
        CantidadProvTROZA = request.form['CantidadProvTROZA']

        data = {
            'loteTROZA': loteTROZA,
            'fecha_ingresoTROZA': fecha_ingresoTROZA,
            'proveedorTROZA': proveedorTROZA,
            'receptorTROZA': receptorTROZA,
            'remision_provTROZA': remision_provTROZA,
            'talo_salvoconductoTROZA': talo_salvoconductoTROZA,
            'remi_salvoconductoTROZA': remi_salvoconductoTROZA,
            'nombre_razonTROZA': nombre_razonTROZA,
            'entidadTROZA': entidadTROZA,
            'no_registroTROZA': no_registroTROZA,
            'departamentoTROZA': departamentoTROZA,
            'municipioTROZA': municipioTROZA,
            'veredaTROZA': veredaTROZA,
            'transportadorTROZA': transportadorTROZA,
            'placaTROZA': placaTROZA,
            'comunTROZA': comunTROZA,
            'cientificoTROZA': cientificoTROZA,
            'presentacionTROZA': presentacionTROZA,
            'infotabletroza': infotabletroza,
            'volm3TROZA': volm3TROZA,
            'M3ProvTROZA': M3ProvTROZA,
            'CantidadProvTROZA': CantidadProvTROZA,
        }

        response = requests.post(
            'http://oficial2024.maderasindustrialesoftware.com/info_entrada/informeTROZA.php', data=data)

        if response.status_code == 200:
            return redirect(url_for('info_entrada'))
        else:
            flash("Error al crear el informe.")
    else:
        # Renderizar el formulario con el número de lote siguiente
        return render_template('info_entrada_html/agregar_infoTROZA.html', lote=generate_lote)


@app.route('/ver_infoPP')
@login_required
def ver_infoPP():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')

    return render_template('info_entrada_html/ver_infoPP.html')


@app.route('/ver_infoBLOQUE')
@login_required
def ver_infoBLOQUE():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')

    return render_template('info_entrada_html/ver_infoBLOQUE.html')


@app.route('/ver_infoTROZA')
@login_required
def ver_infoTROZA():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')

    return render_template('info_entrada_html/ver_infoTROZA.html')


@app.route('/info_prov_val')
@login_required
def info_prov_val():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')

    return render_template('info_entrada_html/info_prov_val.html')


@app.route('/info_entrada_val')
@login_required
def info_entrada_val():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    model_infoPP_response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/get_infoPP.php')
    model_infoPP = model_infoPP_response.json()

    model_infoBLOQUE = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/get_infoBLOQUE.php')
    model_infoBLOQUE = model_infoBLOQUE.json()

    model_infoTROZA = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/get_infoTROZA.php')
    model_infoTROZA = model_infoTROZA.json()

    validado_infoPP = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/get_info_valPP.php')
    validado_infoPP = validado_infoPP.json()

    validado_infoBLOQUE = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/get_info_valBLOQUE.php')
    validado_infoBLOQUE = validado_infoBLOQUE.json()

    validado_infoTROZA = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/get_info_valTROZA.php')
    validado_infoTROZA = validado_infoTROZA.json()

    response = requests.get(
        'http://oficial2024.maderasindustrialesoftware.com/info_entrada/last_lote.php')
    lote = response.text.strip()

    merged_users = merge_data(model_infoPP, model_infoBLOQUE, model_infoTROZA,
                              validado_infoPP, validado_infoBLOQUE, validado_infoTROZA)

    return render_template('info_entrada_html/info_entrada_val.html', users=merged_users, last_infoLOTE=lote)


def merge_data(model_infoPP, model_infoBLOQUE, model_infoTROZA, validado_infoPP, validado_infoBLOQUE, validado_infoTROZA):
    merged_users = []

    for model_userPP in model_infoPP:
        merged_user = {
            'lote': model_userPP['lotePP'],
            'fecha_ingreso': model_userPP['fecha_ingresoPP'],
            'proveedor': model_userPP['proveedorPP'],
            'presentacion': model_userPP['presentacionPP'],
            'm3PP': model_userPP['m3PP'],
            'pulgPP': model_userPP['pulgPP'],
            'cantidadtotalPP': model_userPP['cantidadtotalPP'],
            'infotablepp': model_userPP['infotablepp'],

        }
        merged_users.append(merged_user)

        for validado_user in validado_infoPP:
            if model_userPP['lotePP'] == validado_user['id_lotePP']:
                merged_user['pulg_totalPP'] = validado_user['pulg_totalPP']
                merged_user['m3_totalPP'] = validado_user['m3_totalPP']
                merged_user['fecha'] = validado_user['fechaPP']
                merged_user['mp_gorgojo'] = validado_user['mp_gorgojoPP']
                merged_user['sino_gorgojo'] = validado_user['sino_gorgojoPP']
                merged_user['mp_partida'] = validado_user['mp_partidaPP']
                merged_user['sino_partida'] = validado_user['sino_partidaPP']
                merged_user['mp_fisicos'] = validado_user['mp_fisicosPP']
                merged_user['sino_fisicos'] = validado_user['sino_fisicosPP']
                merged_user['mp_quimicos'] = validado_user['mp_quimicosPP']
                merged_user['sino_quimicos'] = validado_user['sino_quimicosPP']
                merged_user['mp_biologico'] = validado_user['mp_biologicoPP']
                merged_user['sino_biologico'] = validado_user['sino_biologicoPP']
                merged_user['unidades_totalPP'] = validado_user['unidades_totalPP']
                merged_user['mp_validado_porPP'] = validado_user['mp_validado_porPP']
                merged_user['verificado'] = validado_user['verificado']
                merged_user['cantidadesxlargoPP'] = validado_user['cantidadesxlargoPP']
                merged_user['url_img'] = validado_user['url_img']

                break

    for model_userBLOQUE in model_infoBLOQUE:
        merged_user = {
            'lote': model_userBLOQUE['loteBLOQUE'],
            'fecha_ingreso': model_userBLOQUE['fecha_ingresoBLOQUE'],
            'proveedor': model_userBLOQUE['proveedorBLOQUE'],
            'presentacion': model_userBLOQUE['presentacionBLOQUE'],
            'pulgadasBLOQUE': model_userBLOQUE['pulgadasBLOQUE'],
            'infotablebloque': model_userBLOQUE['infotablebloque'],
        }
        merged_users.append(merged_user)

        for validado_user in validado_infoBLOQUE:
            if model_userBLOQUE['loteBLOQUE'] == validado_user['id_loteBLOQUE']:
                merged_user['pulg_totalBLOQUE'] = validado_user['pulg_totalBLOQUE']
                merged_user['m3_totalBLOQUE'] = validado_user['m3_totalBLOQUE']
                merged_user['fecha'] = validado_user['fechaBLOQUE']
                merged_user['mp_validado_por'] = validado_user['mp_validado_porBLOQUE']
                merged_user['mp_gorgojo'] = validado_user['mp_gorgojoBLOQUE']
                merged_user['sino_gorgojo'] = validado_user['sino_gorgojoBLOQUE']
                merged_user['mp_partida'] = validado_user['mp_partidaBLOQUE']
                merged_user['sino_partida'] = validado_user['sino_partidaBLOQUE']
                merged_user['mp_fisicos'] = validado_user['mp_fisicosBLOQUE']
                merged_user['sino_fisicos'] = validado_user['sino_fisicosBLOQUE']
                merged_user['mp_quimicos'] = validado_user['mp_quimicosBLOQUE']
                merged_user['sino_quimicos'] = validado_user['sino_quimicosBLOQUE']
                merged_user['mp_biologico'] = validado_user['mp_biologicoBLOQUE']
                merged_user['sino_biologico'] = validado_user['sino_biologicoBLOQUE']
                merged_user['unidades_totalBLOQUE'] = validado_user['unidades_totalBLOQUE']
                merged_user['verificado'] = validado_user['verificado']
                merged_user['lado1BLOQUE'] = validado_user['lado1BLOQUE']
                merged_user['lado2BLOQUE'] = validado_user['lado2BLOQUE']
                merged_user['lengthBLOQUE'] = validado_user['lengthBLOQUE']
                merged_user['cantidadesxlargoBLOQUE'] = validado_user['cantidadesxlargoBLOQUE']
                merged_user['url_img'] = validado_user['url_img']

                break

    for model_userTROZA in model_infoTROZA:
        merged_user = {
            'lote': model_userTROZA['loteTROZA'],
            'fecha_ingreso': model_userTROZA['fecha_ingresoTROZA'],
            'proveedor': model_userTROZA['proveedorTROZA'],
            'presentacion': model_userTROZA['presentacionTROZA'],
            'volm3TROZA': model_userTROZA['volm3TROZA'],
            'M3Prov': model_userTROZA['M3ProvTROZA'],
            'CantidadProvTROZA': model_userTROZA['CantidadProvTROZA'],
            'infotabletroza': model_userTROZA['infotabletroza'],

        }

        # Found a match in PP data, no need to continue searching

        # Found a match in BLOQUE data, no need to continue searching

        for validado_user in validado_infoTROZA:
            if model_userTROZA['loteTROZA'] == validado_user['id_loteTROZA']:
                merged_user['vol_m3_totalTROZA'] = validado_user['vol_m3_totalTROZA']
                merged_user['fecha'] = validado_user['fechaTROZA']
                merged_user['mp_validado_por'] = validado_user['mp_validado_porTROZA']
                merged_user['mp_gorgojo'] = validado_user['mp_gorgojoTROZA']
                merged_user['sino_gorgojo'] = validado_user['sino_gorgojoTROZA']
                merged_user['mp_partida'] = validado_user['mp_partidaTROZA']
                merged_user['sino_partida'] = validado_user['sino_partidaTROZA']
                merged_user['mp_fisicos'] = validado_user['mp_fisicosTROZA']
                merged_user['sino_fisicos'] = validado_user['sino_fisicosTROZA']
                merged_user['mp_quimicos'] = validado_user['mp_quimicosTROZA']
                merged_user['sino_quimicos'] = validado_user['sino_quimicosTROZA']
                merged_user['mp_biologico'] = validado_user['mp_biologicoTROZA']
                merged_user['sino_biologico'] = validado_user['sino_biologicoTROZA']
                merged_user['unidades_totalTROZA'] = validado_user['unidades_totalTROZA']
                merged_user['puntaTROZA'] = validado_user['puntaTROZA']
                merged_user['lado1TROZA'] = validado_user['lado1TROZA']
                merged_user['lado2TROZA'] = validado_user['lado2TROZA']
                merged_user['lengthTROZA'] = validado_user['lengthTROZA']
                merged_user['verificado'] = validado_user['verificado']
                merged_user['cantidadesxlargoTROZA'] = validado_user['cantidadesxlargoTROZA']
                merged_user['url_img'] = validado_user['url_img']
                break

        merged_users.append(merged_user)

    return merged_users


@app.route('/ver_prov_val')
@login_required
def ver_prov_val():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    return render_template('info_entrada_html/ver_prov_val.html')


@app.route('/ver_prov_valPP')
@login_required
def ver_prov_valPP():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    return render_template('info_entrada_html/ver_prov_valPP.html')


@app.route('/update_statusPP', methods=['POST'])
def update_statusPP():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    try:
        if request.method == 'POST':
            # Obtener datos del formulario
            lote = request.form['lote']
            status = request.form['status']

            data = {
                'lote': lote,
                'status': status,
            }

            response = requests.post(
                'http://oficial2024.maderasindustrialesoftware.com/info_entrada/verificacionPP.php', data=data)

            return jsonify({'message': 'Status and lote received successfully.'}), 200

    except Exception as ex:
        # Manejar cualquier excepción que pueda ocurrir durante el procesamiento
        return jsonify({'error': str(ex)}), 500


@app.route('/ver_prov_valBLOQUE')
@login_required
def ver_prov_valBLOQUE():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    return render_template('info_entrada_html/ver_prov_valBLOQUE.html')


@app.route('/update_statusBLOQUE', methods=['POST'])
def update_statusBLOQUE():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    try:
        if request.method == 'POST':
            # Obtener datos del formulario
            lote = request.form['lote']
            status = request.form['status']

            data = {
                'lote': lote,
                'status': status,
            }

            response = requests.post(
                'http://oficial2024.maderasindustrialesoftware.com/info_entrada/verificacionBLOQUE.php', data=data)

            return jsonify({'message': 'Status and lote received successfully.'}), 200

    except Exception as ex:
        # Manejar cualquier excepción que pueda ocurrir durante el procesamiento
        return jsonify({'error': str(ex)}), 500


@app.route('/ver_prov_valTROZA')
@login_required
def ver_prov_valTROZA():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')

    return render_template('info_entrada_html/ver_prov_valTROZA.html')


@app.route('/update_statusTROZA', methods=['POST'])
def update_statusTROZA():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    try:
        if request.method == 'POST':
            # Obtener datos del formulario
            lote = request.form['lote']
            status = request.form['status']

            data = {
                'lote': lote,
                'status': status,
            }

            response = requests.post(
                'http://oficial2024.maderasindustrialesoftware.com/info_entrada/verificacionTROZA.php', data=data)

            return jsonify({'message': 'Status and lote received successfully.'}), 200

    except Exception as ex:
        # Manejar cualquier excepción que pueda ocurrir durante el procesamiento
        return jsonify({'error': str(ex)}), 500
    

#*********ELIMINAR INFORME LIBERADO*********

@app.route('/EliminarPPVal', methods=['POST'])
@login_required
def EliminarPPVal():
    try:
        lote = request.json['lote']
        response = requests.post(
            'http://oficial2024.maderasindustrialesoftware.com/info_entrada/Eliminar_LotePP.php', data={'lote': lote})
        return jsonify({'message': response.json()['message']}), response.status_code
    except Exception as ex:
        return jsonify({'error': str(ex)}), 500

    
@app.route('/EliminarBloqueVal', methods=['POST'])
@login_required
def EliminarBloqueVal():
    try:
        lote = request.json['lote']
        response = requests.post(
            'http://oficial2024.maderasindustrialesoftware.com/info_entrada/Eliminar_LoteBLOQUE.php', data={'lote': lote})
        return jsonify({'message': response.json()['message']}), response.status_code
    except Exception as ex:
        return jsonify({'error': str(ex)}), 500
    
@app.route('/EliminarTrozaVal', methods=['POST'])
@login_required
def EliminarTrozaVal():
    try:
        lote = request.json['lote']
        response = requests.post(
            'http://oficial2024.maderasindustrialesoftware.com/info_entrada/Eliminar_LoteTROZA.php', data={'lote': lote})
        return jsonify({'message': response.json()['message']}), response.status_code
    except Exception as ex:
        return jsonify({'error': str(ex)}), 500


@app.route('/perfil')
@login_required
def perfil():
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')

    return render_template('admin_html/perfil.html')


@app.route('/error_conection_mi_software')
def error_conection_mi_software():
    if app.config['maintenance_mode']:
        # Mantenimiento habilitado, muestra la página de mantenimiento
        return """
    <html style="user-select: none; background-color: #fff;">
    <head>
        <style>
            body {
                font-family: 'Arial', sans-serif;
                background-color: #fff;
            }
            .container {
                text-align: center;
                padding: 22px;
                margin-top: 100px; /* Add a top margin to the container */
            }
            h1 {
                font-family: 'Georgia', serif;
                font-size: 29px;
                color: #333;
            }
            p {
                font-family: 'Georgia', serif;
                font-size: 24px;
                color: #666;
            }
            em {
                font-style: italic;
            }
            img {
                max-width: 100%;
                height: auto;
                display: block;
                margin: 0 auto;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>Ops, Algo salió mal...</h1>
            <h1>🛠️Actualmente estamos realizando mantenimiento. Por favor, vuelve más tarde. 🕒</h1>
            <img src="../static/img/logo.png">
            <p><em>Maderas Industriales S.A.S</em></p>
        </div>
    </body>
    </html>
    """
    else:
        # Mantenimiento deshabilitado, redirige a la página de inicio de sesión
        return redirect(url_for('home'))


@app.errorhandler(requests.exceptions.ConnectionError)
def handle_connection_error(e):
    if app.config['maintenance_mode']:
        return redirect('/error_conection_mi_software')
    else:
        # Continúa manejando el error de otra manera, si es necesario
        pass


@app.route('/protected')
@login_required
def protected():
    return "<h1>Esta es una vista protegida, solo para usuarios autenticados.</h1>"


def status_401(error):
    return redirect(url_for('login'))


def status_404(error):
    return "<h1>Página no encontrada</h1>", 404


if __name__ == '__main__':
    app.config.from_object(config['development'])
    csrf.init_app(app)
    app.register_error_handler(401, status_401)
    app.register_error_handler(404, status_404)
    app.run()
